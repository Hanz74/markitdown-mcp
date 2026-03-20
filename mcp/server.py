#!/usr/bin/env python3
# Aktiviere uvloop für bessere Performance
try:
    import uvloop
    uvloop.install()
except ImportError:
    pass

"""
MarkItDown MCP + REST Server v0.3.0

Bietet zwei Schnittstellen:
- MCP (Port 8080): Für Claude und andere MCP-Clients
- REST (Port 8081): Für n8n und andere HTTP-Clients

Features:
- Auto-Routing: Bilder → Vision, Dokumente → MarkItDown
- Bild-Resize vor Vision (spart Tokens)
- Folder-Konvertierung (alle Dateien in einem Ordner)
- URL-Konvertierung
- Retry-Logik für API-Calls
- Strukturiertes Logging
"""

import os
import io
import json
import base64
import hashlib
import mimetypes
import re
import threading
import time
import logging
import zipfile
from datetime import datetime
from typing import Optional, Any
from pathlib import Path

import subprocess

import httpx
import uvicorn
import magic
import structlog
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False
from PIL import Image
from fastapi import FastAPI, HTTPException, Request
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from fastmcp import FastMCP
from markitdown import MarkItDown
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

try:
    import jsonschema
    JSONSCHEMA_AVAILABLE = True
except ImportError:
    JSONSCHEMA_AVAILABLE = False

try:
    from faster_whisper import WhisperModel
    WHISPER_AVAILABLE = True
except ImportError:
    WHISPER_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from img2table.document import PDF as Img2TablePDF
    from img2table.ocr import TesseractOCR
    IMG2TABLE_AVAILABLE = True
except ImportError:
    IMG2TABLE_AVAILABLE = False

from models import (
    ConvertRequest,
    ConvertResponse,
    ConvertFolderRequest,
    AnalyzeRequest,
    ExtractRequest,
    TemplateResponse,
    HealthResponse,
    MetaData,
    ErrorCode,
    create_error_response,
    create_success_response,
)


# =============================================================================
# Konfiguration
# =============================================================================

DATA_DIR = Path(os.getenv("DATA_DIR", "/data"))
TEMP_DIR = Path(os.getenv("TEMP_DIR", "/tmp/markitdown"))
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# Mistral Vision
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY", "")
MISTRAL_API_URL = os.getenv("MISTRAL_API_URL", "https://api.mistral.ai/v1")
MISTRAL_VISION_MODEL = os.getenv("MISTRAL_VISION_MODEL", "mistral-small-2603")
MISTRAL_TIMEOUT = int(os.getenv("MISTRAL_TIMEOUT", "120"))

# Mistral OCR 3
MISTRAL_OCR_MODEL = os.getenv("MISTRAL_OCR_MODEL", "mistral-ocr-2512")
MISTRAL_OCR_ENABLED = os.getenv("MISTRAL_OCR_ENABLED", "true").lower() == "true"

# Server Ports
MCP_PORT = int(os.getenv("MCP_PORT", "8080"))
REST_PORT = int(os.getenv("REST_PORT", "8081"))

# Limits
MAX_FILE_SIZE_MB = int(os.getenv("MAX_FILE_SIZE_MB", "25"))
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
IMAGE_MAX_WIDTH = int(os.getenv("IMAGE_MAX_WIDTH", "2048"))

# Retry
MAX_RETRIES = int(os.getenv("MAX_RETRIES", "3"))

# Scanned PDF Detection
SCAN_THRESHOLD_CHARS = int(os.getenv("SCAN_THRESHOLD_CHARS", "50"))

# Klassifizierung
DEFAULT_CLASSIFY_CATEGORIES = os.getenv(
    "CLASSIFY_CATEGORIES",
    "invoice,contract,cv,protocol,letter,technical_doc,report,presentation,spreadsheet,other"
).split(",")

# Logging
LOG_LEVEL = os.getenv("LOG_LEVEL", "info").upper()
LOG_FORMAT = os.getenv("LOG_FORMAT", "json")

# Dateitypen
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
MARKITDOWN_EXTENSIONS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls",
                          ".odt", ".ods", ".odp", ".html", ".htm", ".xml", ".json",
                          ".csv", ".txt", ".md", ".rtf"}
SKIP_FILES = set(os.getenv("SKIP_FILES", "email.md,consolidated.md,metadata.json,.DS_Store,Thumbs.db").split(","))

# Audio/Video (FR-MKIT-006)
WHISPER_MODEL_SIZE = os.getenv("WHISPER_MODEL_SIZE", "base")
AUDIO_EXTENSIONS = {".mp3", ".wav", ".ogg", ".flac", ".m4a"}
VIDEO_EXTENSIONS = {".mp4", ".mkv", ".webm", ".avi", ".mov"}

# Modelle — separate Modelle für verschiedene Tasks
MISTRAL_TEXT_MODEL = os.getenv("MISTRAL_TEXT_MODEL", "mistral-small-2603")

# Token-Limits — großzügig, Kosten nur pro verbrauchtem Token
VISION_MAX_TOKENS = int(os.getenv("VISION_MAX_TOKENS", "16384"))
CLASSIFY_MAX_TOKENS = int(os.getenv("CLASSIFY_MAX_TOKENS", "1024"))
EXTRACT_MAX_TOKENS = int(os.getenv("EXTRACT_MAX_TOKENS", "16384"))
OCR_CORRECT_MAX_TOKENS = int(os.getenv("OCR_CORRECT_MAX_TOKENS", "16384"))

# Text-Limits für LLM-Input
CLASSIFY_MAX_CHARS = int(os.getenv("CLASSIFY_MAX_CHARS", "32000"))
EXTRACT_MAX_CHARS = int(os.getenv("EXTRACT_MAX_CHARS", "32000"))

# Sprache
DEFAULT_LANGUAGE = os.getenv("DEFAULT_LANGUAGE", "de")

# Bild-Verarbeitung
MIN_IMAGE_SIZE_PX = int(os.getenv("MIN_IMAGE_SIZE_PX", "50"))
PDF_RENDER_DPI = int(os.getenv("PDF_RENDER_DPI", "200"))

# Timeouts
PDFTOTEXT_TIMEOUT = int(os.getenv("PDFTOTEXT_TIMEOUT", "60"))
PDFINFO_TIMEOUT = int(os.getenv("PDFINFO_TIMEOUT", "30"))
FFMPEG_TIMEOUT = int(os.getenv("FFMPEG_TIMEOUT", "600"))

# Whisper
WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cpu")
WHISPER_COMPUTE_TYPE = os.getenv("WHISPER_COMPUTE_TYPE", "int8")

# Whisper-Modell-Cache (wird beim ersten Aufruf geladen)
_whisper_model_cache: dict[str, Any] = {}  # key: model_size → WhisperModel instance

VERSION = "0.3.0"
START_TIME = time.time()


# =============================================================================
# Logging Setup
# =============================================================================

def setup_logging():
    """Konfiguriert strukturiertes Logging."""
    structlog.configure(
        processors=[
            structlog.stdlib.filter_by_level,
            structlog.stdlib.add_logger_name,
            structlog.stdlib.add_log_level,
            structlog.processors.TimeStamper(fmt="iso"),
            structlog.processors.StackInfoRenderer(),
            structlog.processors.format_exc_info,
            structlog.processors.UnicodeDecoder(),
            structlog.processors.JSONRenderer() if LOG_FORMAT == "json" else structlog.dev.ConsoleRenderer(),
        ],
        context_class=dict,
        logger_factory=structlog.stdlib.LoggerFactory(),
        wrapper_class=structlog.stdlib.BoundLogger,
        cache_logger_on_first_use=True,
    )
    logging.basicConfig(
        format="%(message)s",
        level=getattr(logging, LOG_LEVEL, logging.INFO),
    )

setup_logging()
log = structlog.get_logger()


# =============================================================================
# Initialisierung
# =============================================================================

# MarkItDown Instanz
md = MarkItDown()

# FastMCP für MCP-Protokoll
mcp = FastMCP("markitdown_mcp")

# FastAPI für REST
app = FastAPI(
    title="MarkItDown API",
    description="Konvertiert Dokumente und Bilder zu Markdown",
    version=VERSION,
    docs_url="/docs",
    redoc_url="/redoc",
)


def _safe_encode(obj: Any) -> Any:
    """Recursively converts bytes values to a safe string representation."""
    if isinstance(obj, bytes):
        try:
            return obj.decode("utf-8")
        except UnicodeDecodeError:
            return f"<binary {len(obj)} bytes>"
    if isinstance(obj, dict):
        return {k: _safe_encode(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_safe_encode(v) for v in obj]
    return obj


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError) -> JSONResponse:
    """
    Custom handler for request validation errors.

    FastAPI's default handler uses jsonable_encoder which calls bytes.decode(),
    causing UnicodeDecodeError when binary data (e.g. PDF uploads via multipart)
    appears in the error details.
    """
    safe_errors = _safe_encode(exc.errors())
    return JSONResponse(
        status_code=422,
        content={"detail": safe_errors},
    )


# =============================================================================
# Hilfsfunktionen
# =============================================================================

def resolve_path(path: str) -> Path:
    """Löst relativen Pfad zu absolutem auf."""
    p = Path(path)
    if p.is_absolute():
        return p
    return DATA_DIR / p


def get_file_extension(filename: str) -> str:
    """Extrahiert die Dateiendung."""
    return Path(filename).suffix.lower()


def is_image_file(path: Path) -> bool:
    """Prüft ob eine Datei ein unterstütztes Bild ist."""
    return path.suffix.lower() in IMAGE_EXTENSIONS


def is_markitdown_file(path: Path) -> bool:
    """Prüft ob eine Datei von MarkItDown verarbeitet werden kann."""
    return path.suffix.lower() in MARKITDOWN_EXTENSIONS


def is_audio_file(path: Path) -> bool:
    """Prüft ob eine Datei eine unterstützte Audio-Datei ist (FR-MKIT-006).

    Args:
        path: Pfad zur Datei.

    Returns:
        True wenn die Dateiendung in AUDIO_EXTENSIONS enthalten ist.
    """
    return path.suffix.lower() in AUDIO_EXTENSIONS


def is_video_file(path: Path) -> bool:
    """Prüft ob eine Datei eine unterstützte Video-Datei ist (FR-MKIT-006).

    Args:
        path: Pfad zur Datei.

    Returns:
        True wenn die Dateiendung in VIDEO_EXTENSIONS enthalten ist.
    """
    return path.suffix.lower() in VIDEO_EXTENSIONS


def extract_audio_from_video(video_path: Path) -> Path:
    """Extrahiert den Audio-Track aus einer Video-Datei als WAV via ffmpeg (FR-MKIT-006).

    Args:
        video_path: Pfad zur Video-Datei.

    Returns:
        Pfad zur extrahierten WAV-Datei im TEMP_DIR.

    Raises:
        RuntimeError: Wenn ffmpeg fehlschlägt oder nicht installiert ist.
    """
    wav_filename = f"{video_path.stem}_{hashlib.md5(str(video_path).encode()).hexdigest()[:8]}.wav"
    wav_path = TEMP_DIR / wav_filename

    log.info("extract_audio_from_video_start", video=str(video_path), output=str(wav_path))

    try:
        result = subprocess.run(
            [
                "ffmpeg", "-y",
                "-i", str(video_path),
                "-vn",               # kein Video
                "-acodec", "pcm_s16le",
                "-ar", "16000",      # 16kHz — optimal für Whisper
                "-ac", "1",          # Mono
                str(wav_path),
            ],
            capture_output=True,
            text=True,
            timeout=FFMPEG_TIMEOUT,
        )
        if result.returncode != 0:
            log.error(
                "extract_audio_from_video_failed",
                video=str(video_path),
                returncode=result.returncode,
                stderr=result.stderr,
            )
            raise RuntimeError(f"ffmpeg fehlgeschlagen (returncode={result.returncode}): {result.stderr}")
    except FileNotFoundError as exc:
        raise RuntimeError("ffmpeg ist nicht installiert") from exc
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError("ffmpeg-Timeout beim Audio-Extrahieren") from exc

    log.info("extract_audio_from_video_done", wav=str(wav_path))
    return wav_path


def transcribe_audio(audio_path: Path) -> dict[str, Any]:
    """Transkribiert eine Audio-Datei mit faster-whisper (FR-MKIT-006).

    Nutzt WHISPER_MODEL_SIZE aus der Umgebungsvariable.
    Das Modell wird gecacht — nach dem ersten Laden wird es wiederverwendet.

    Args:
        audio_path: Pfad zur Audio-Datei (WAV bevorzugt).

    Returns:
        Dict mit:
        - success (bool)
        - text (str): Vollständiges Transkript
        - language (str): Erkannte Sprache (z.B. "de", "en")
        - duration (float): Dauer in Sekunden
        - model_size (str): Verwendete Modell-Größe
        - error (str, optional): Fehlermeldung bei Misserfolg
    """
    if not WHISPER_AVAILABLE:
        log.warning("transcribe_audio_whisper_not_available", file=str(audio_path))
        return {
            "success": False,
            "error": "faster-whisper ist nicht installiert (pip install faster-whisper)",
        }

    log.info("transcribe_audio_start", file=str(audio_path), model_size=WHISPER_MODEL_SIZE)

    try:
        # Modell aus Cache laden oder frisch initialisieren
        if WHISPER_MODEL_SIZE not in _whisper_model_cache:
            log.info("whisper_model_load", model_size=WHISPER_MODEL_SIZE)
            _whisper_model_cache[WHISPER_MODEL_SIZE] = WhisperModel(
                WHISPER_MODEL_SIZE,
                device=WHISPER_DEVICE,
                compute_type=WHISPER_COMPUTE_TYPE,
            )
        model = _whisper_model_cache[WHISPER_MODEL_SIZE]

        segments, info = model.transcribe(str(audio_path), beam_size=5)

        # Segmente zusammenführen
        text_parts: list[str] = []
        duration = 0.0
        for segment in segments:
            text_parts.append(segment.text.strip())
            duration = max(duration, segment.end)

        full_text = " ".join(text_parts)
        detected_language = info.language if hasattr(info, "language") else "unknown"

        log.info(
            "transcribe_audio_done",
            file=str(audio_path),
            language=detected_language,
            duration=duration,
            chars=len(full_text),
        )

        return {
            "success": True,
            "text": full_text,
            "language": detected_language,
            "duration": duration,
            "model_size": WHISPER_MODEL_SIZE,
        }

    except Exception as exc:
        log.error("transcribe_audio_error", file=str(audio_path), error=str(exc))
        return {
            "success": False,
            "error": f"Transkription fehlgeschlagen: {str(exc)}",
        }


def should_skip_file(filename: str) -> bool:
    """Prüft ob eine Datei übersprungen werden soll."""
    return filename in SKIP_FILES or filename.startswith(".")


def get_mimetype(path: Path) -> str:
    """Ermittelt den MIME-Type einer Datei."""
    suffix = path.suffix.lower()
    mime_map = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
        ".pdf": "application/pdf",
    }
    return mime_map.get(suffix, "application/octet-stream")


def detect_mimetype_from_bytes(data: bytes) -> Optional[str]:
    """Erkennt MIME-Type aus Magic Bytes."""
    try:
        return magic.from_buffer(data, mime=True)
    except Exception:
        return None


def resize_image_if_needed(image_data: bytes, max_width: int = IMAGE_MAX_WIDTH) -> tuple[bytes, dict]:
    """
    Verkleinert ein Bild falls es zu groß ist.

    Returns:
        tuple: (image_bytes, resize_meta)
    """
    resize_meta = {
        "resized": False,
        "original_width": None,
        "original_height": None,
    }

    try:
        img = Image.open(io.BytesIO(image_data))
        resize_meta["original_width"] = img.width
        resize_meta["original_height"] = img.height

        if img.width > max_width:
            ratio = max_width / img.width
            new_height = int(img.height * ratio)
            img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
            resize_meta["resized"] = True
            resize_meta["width"] = max_width
            resize_meta["height"] = new_height

            output = io.BytesIO()
            img_format = "JPEG" if img.mode == "RGB" else "PNG"
            if img.mode == "RGBA" and img_format == "JPEG":
                img = img.convert("RGB")
            img.save(output, format=img_format, quality=85)
            return output.getvalue(), resize_meta
        else:
            resize_meta["width"] = img.width
            resize_meta["height"] = img.height
            return image_data, resize_meta

    except Exception as e:
        log.warning("image_resize_failed", error=str(e))
        return image_data, resize_meta


# =============================================================================
# LLM Artifact Cleanup (T-MKIT-016)
# =============================================================================

# Patterns for LLM preambles that should be stripped (German and English)
_PREAMBLE_PATTERNS = [
    # German patterns
    re.compile(
        r'^(?:hier\s+ist\s+(?:der|die|das)\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:im\s+folgenden\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:nachfolgend\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:gerne[,!.]?\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    # English patterns
    re.compile(
        r'^(?:here\s+is\s+(?:the\s+)?[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:below\s+is\s+(?:the\s+)?[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:the\s+following\s+[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:certainly[!,.]?\s*[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
    re.compile(
        r'^(?:sure[!,.]?\s*[^\n]{0,120}\n+)',
        re.IGNORECASE,
    ),
]

# Pattern for full-output code block wrapping (```markdown ... ``` or ``` ... ```)
_FULL_CODEBLOCK_PATTERN = re.compile(
    r'^```(?:markdown|md)?\s*\n([\s\S]*?)\n```\s*$',
    re.IGNORECASE,
)


def strip_llm_artifacts(text: str) -> str:
    """
    Entfernt typische LLM-Artefakte aus dem Output.

    Entfernt:
    - Einleitende Preamble-Sätze ("Hier ist...", "Here is...", "Im Folgenden...", etc.)
    - ```markdown ... ``` Wrapping wenn der gesamte Output darin eingeschlossen ist
    - ``` ... ``` Wrapping wenn der gesamte Output darin eingeschlossen ist

    Codeblöcke INNERHALB des Textes werden NICHT entfernt.

    Args:
        text: LLM-Output-Text der bereinigt werden soll.

    Returns:
        Bereinigter Text.
    """
    if not text:
        return text

    result = text.strip()

    # Mermaid-Codeblöcke NICHT strippen — sie sind gewollter Output
    if result.startswith("```mermaid"):
        return result

    # 1. Outer code block wrapping entfernen (```markdown ... ``` oder ``` ... ```)
    codeblock_match = _FULL_CODEBLOCK_PATTERN.match(result)
    if codeblock_match:
        result = codeblock_match.group(1).strip()

    # 2. Preamble-Zeilen am Anfang entfernen
    for pattern in _PREAMBLE_PATTERNS:
        result = pattern.sub('', result)
        result = result.strip()

    return result


# =============================================================================
# API Calls mit Retry
# =============================================================================

@retry(
    stop=stop_after_attempt(MAX_RETRIES),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.HTTPStatusError)),
    reraise=True,
)
async def call_mistral_vision_api(payload: dict) -> dict:
    """Ruft die Mistral Vision API mit Retry-Logik auf."""
    async with httpx.AsyncClient(timeout=float(MISTRAL_TIMEOUT)) as client:
        response = await client.post(
            f"{MISTRAL_API_URL}/chat/completions",
            headers={
                "Authorization": f"Bearer {MISTRAL_API_KEY}",
                "Content-Type": "application/json"
            },
            json=payload
        )
        response.raise_for_status()
        return response.json()


@retry(
    stop=stop_after_attempt(MAX_RETRIES),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.HTTPStatusError)),
    reraise=True,
)
async def call_mistral_ocr_api(file_data: bytes, filename: str) -> dict:
    """Ruft die Mistral OCR API (/v1/ocr) auf."""
    b64 = base64.b64encode(file_data).decode("utf-8")
    payload = {
        "model": MISTRAL_OCR_MODEL,
        "document": {
            "type": "document_url",
            "document_url": f"data:application/pdf;base64,{b64}",
        },
    }
    async with httpx.AsyncClient(timeout=float(MISTRAL_TIMEOUT)) as client:
        response = await client.post(
            f"{MISTRAL_API_URL}/ocr",
            headers={
                "Authorization": f"Bearer {MISTRAL_API_KEY}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
        response.raise_for_status()
        return response.json()


async def analyze_with_mistral_vision(
    image_data: bytes,
    mimetype: str,
    prompt: str,
    language: str = "de"
) -> dict[str, Any]:
    """Analysiert ein Bild mit Mistral Pixtral Vision API."""
    if not MISTRAL_API_KEY:
        return {
            "success": False,
            "error_code": ErrorCode.API_KEY_INVALID,
            "error": "MISTRAL_API_KEY nicht konfiguriert"
        }

    b64_image = base64.b64encode(image_data).decode("utf-8")
    data_url = f"data:{mimetype};base64,{b64_image}"

    system_prompt = (
        "Du bist ein präziser Assistent für Bild- und Dokumentenanalyse. "
        "Befolge ausschließlich die Anweisungen im User-Prompt. "
        "Antworte NICHT mit Einleitungssätzen, Erklärungen oder Code-Block-Wrapping — nur mit dem angeforderten Ergebnis."
        if language == "de"
        else "You are a precise assistant for image and document analysis. "
        "Follow only the instructions in the user prompt. "
        "Do NOT reply with introductions, explanations, or code-block wrapping — only the requested result."
    )

    payload = {
        "model": MISTRAL_VISION_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": data_url}}
                ]
            }
        ],
        "max_tokens": VISION_MAX_TOKENS
    }

    try:
        log.info("vision_api_call", model=MISTRAL_VISION_MODEL, image_size=len(image_data))
        result = await call_mistral_vision_api(payload)

        content = result["choices"][0]["message"]["content"]
        content = strip_llm_artifacts(content)
        usage = result.get("usage", {})

        log.info("vision_api_success", tokens=usage.get("total_tokens", 0))
        return {
            "success": True,
            "markdown": content,
            "tokens_prompt": usage.get("prompt_tokens", 0),
            "tokens_completion": usage.get("completion_tokens", 0),
            "tokens_total": usage.get("total_tokens", 0),
            "vision_model": MISTRAL_VISION_MODEL,
        }

    except httpx.TimeoutException:
        log.error("vision_api_timeout", timeout=MISTRAL_TIMEOUT)
        return {
            "success": False,
            "error_code": ErrorCode.TIMEOUT,
            "error": f"Mistral API Timeout nach {MISTRAL_TIMEOUT}s"
        }
    except httpx.HTTPStatusError as e:
        error_detail = str(e)
        try:
            error_detail = e.response.json().get("error", {}).get("message", str(e))
        except Exception:
            pass
        log.error("vision_api_error", error=error_detail)
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"Mistral API Fehler: {error_detail}"
        }
    except Exception as e:
        log.error("vision_api_exception", error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.VISION_FAILED,
            "error": f"Vision-Analyse fehlgeschlagen: {str(e)}"
        }


async def dual_pass_validate(
    markdown: str,
    file_data: bytes,
    mimetype: str,
    language: str = "de",
) -> str:
    """
    Validiert und korrigiert OCR-extrahierten Markdown via Dual-Pass Vision-Vergleich.

    Schickt den OCR-extrahierten Markdown-Text zusammen mit dem Originalbild an die
    Vision-API. Das Modell vergleicht beides und korrigiert Fehler in Struktur,
    Tabellen-Spalten und Inhalt.

    Bei fehlender Vision-API oder Fehler wird der Original-Markdown zurückgegeben
    (graceful degradation).

    Args:
        markdown: Per OCR extrahierter Markdown-Text.
        file_data: Rohe Bytes der Originaldatei (Bild oder PDF-Seite als Bild).
        mimetype: MIME-Type der Datei (z.B. 'image/png', 'image/jpeg').
        language: Sprache für den Vision-Prompt (Standard: "de").

    Returns:
        Korrigierter Markdown-Text oder Original bei Fehler.
    """
    if not MISTRAL_API_KEY:
        log.warning("dual_pass_validate_skipped_no_api_key")
        return markdown

    prompt = (
        f"Hier ist ein per OCR extrahierter Text und das Originalbild. "
        f"Vergleiche beides und korrigiere Fehler in Struktur, Tabellen-Spalten und Inhalt. "
        f"Gib den korrigierten Markdown zurück. Antworte NUR mit dem korrigierten Text.\n\n"
        f"OCR-Text:\n{markdown}"
    )

    log.info("dual_pass_validate_start", mimetype=mimetype, markdown_len=len(markdown))

    try:
        result = await analyze_with_mistral_vision(
            image_data=file_data,
            mimetype=mimetype,
            prompt=prompt,
            language=language,
        )

        if result.get("success"):
            corrected = result.get("markdown", markdown)
            log.info("dual_pass_validate_done", original_len=len(markdown), corrected_len=len(corrected))
            return corrected
        else:
            log.warning(
                "dual_pass_validate_vision_failed",
                error=result.get("error", "unknown"),
            )
            return markdown

    except Exception as exc:
        log.warning("dual_pass_validate_exception", error=str(exc))
        return markdown


def extract_tables_with_pdfplumber(file_path: Path) -> list[dict]:
    """
    Extrahiert Tabellen aus einer PDF-Datei mit pdfplumber.

    Gibt eine Liste von Seitentabellen zurück, wobei jeder Eintrag Seitennummer
    und die gefundenen Tabellen (als Liste von Zeilen-Listen) enthält.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Liste mit Dictionaries der Form:
        [{"page": int, "tables": list[list[list[str | None]]]}]
        Leere Liste wenn pdfplumber nicht verfügbar oder keine Tabellen gefunden.
    """
    if not PDFPLUMBER_AVAILABLE:
        log.warning("pdfplumber_not_available")
        return []

    page_tables: list[dict] = []
    try:
        with pdfplumber.open(str(file_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                if tables:
                    page_tables.append({"page": page_num, "tables": tables})
        log.info(
            "pdfplumber_extracted",
            file=str(file_path),
            pages_with_tables=len(page_tables),
        )
    except Exception as e:
        log.warning("pdfplumber_error", file=str(file_path), error=str(e))
        return []

    return page_tables


def extract_tables_with_img2table(file_path: Path) -> list[list]:
    """
    Extrahiert Tabellen aus einer gescannten PDF-Datei mit img2table + TesseractOCR.

    Wird als Fallback verwendet wenn pdfplumber keine Tabellen in der Datei findet.
    Nutzt TesseractOCR als OCR-Backend für die Erkennung von Zellinhalten.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Flache Liste von Tabellen (jede Tabelle = Liste von Zeilen, jede Zeile = Liste
        von Strings). Leere Liste wenn img2table nicht verfügbar, keine Tabellen
        gefunden oder ein Fehler auftritt.
    """
    if not IMG2TABLE_AVAILABLE:
        log.warning("img2table_not_available")
        return []

    try:
        ocr = TesseractOCR()
        pdf_doc = Img2TablePDF(src=str(file_path))
        extracted = pdf_doc.extract_tables(ocr=ocr)

        all_tables: list[list] = []
        # extracted ist ein Dict: {page_index: [ExtractedTable, ...]}
        for _page_idx, page_tables in extracted.items():
            for extracted_table in page_tables:
                # ExtractedTable.df ist ein pandas DataFrame
                df = extracted_table.df
                # Erste Zeile als Header, Rest als Datenzeilen
                rows: list[list[str]] = []
                header = [str(col) if col is not None else "" for col in df.columns]
                rows.append(header)
                for _, row in df.iterrows():
                    rows.append([str(v) if v is not None else "" for v in row])
                all_tables.append(rows)

        log.info(
            "img2table_extracted",
            file=str(file_path),
            table_count=len(all_tables),
        )
        return all_tables

    except Exception as e:
        log.warning("img2table_error", file=str(file_path), error=str(e))
        return []


def merge_cross_page_tables(page_tables: list[dict]) -> list[list]:
    """
    Führt Tabellen die über Seitengrenzen hinweg gehen zusammen.

    Algorithmus:
    - Vergleicht die letzte Tabelle auf Seite N mit der ersten Tabelle auf Seite N+1.
    - Wenn beide die gleiche Spaltenanzahl haben: Zusammenführen.
    - Header-Deduplizierung: Wenn die erste Zeile auf der Folgeseite identisch mit
      der ersten Zeile der Ausgangstabelle ist, wird sie weggelassen.
    - Unterschiedliche Spaltenanzahl → separate Tabellen.

    Args:
        page_tables: Ausgabe von extract_tables_with_pdfplumber.

    Returns:
        Flache Liste aller (ggf. zusammengeführten) Tabellen als Zeilen-Listen.
    """
    if not page_tables:
        return []

    # Alle Tabellen mit ihrer Seitenreihenfolge sammeln
    # Jede Tabelle bekommt eine (Seite, Tabellen-Index)-Referenz
    all_tables: list[list] = []

    for page_entry in page_tables:
        for table in page_entry["tables"]:
            all_tables.append(table)

    if not all_tables:
        return []

    # Tabellen iterativ zusammenführen
    merged: list[list] = [all_tables[0]]

    for current_table in all_tables[1:]:
        last_merged = merged[-1]

        if not last_merged or not current_table:
            merged.append(current_table)
            continue

        last_col_count = len(last_merged[0]) if last_merged else 0
        curr_col_count = len(current_table[0]) if current_table else 0

        if last_col_count == curr_col_count and last_col_count > 0:
            # Gleiche Spaltenanzahl → potenzieller Merge
            last_header = last_merged[0]
            curr_header = current_table[0]

            # Header-Deduplizierung: erste Zeile identisch → überspringen
            rows_to_add = current_table
            if curr_header == last_header:
                rows_to_add = current_table[1:]

            if rows_to_add:
                merged[-1] = last_merged + rows_to_add
        else:
            # Unterschiedliche Spaltenanzahl → separate Tabelle
            merged.append(current_table)

    return merged


def tables_to_markdown(tables: list[list]) -> str:
    """
    Konvertiert extrahierte Tabellen in Markdown-Format.

    Jede Tabelle wird als Markdown-Tabelle mit Header-Trennzeile formatiert.
    None-Werte in Zellen werden als leere Strings behandelt.

    Args:
        tables: Liste von Tabellen, jede Tabelle ist eine Liste von Zeilen,
                jede Zeile ist eine Liste von Zellwerten (str | None).

    Returns:
        Zusammengefügter Markdown-String aller Tabellen, getrennt durch Leerzeilen.
    """
    if not tables:
        return ""

    markdown_parts: list[str] = []

    for table in tables:
        if not table:
            continue

        # Zeilen normalisieren: None → leerer String, alle Werte zu str
        normalized_rows: list[list[str]] = []
        for row in table:
            normalized_rows.append([str(cell) if cell is not None else "" for cell in row])

        if not normalized_rows:
            continue

        lines: list[str] = []

        # Header (erste Zeile)
        header = normalized_rows[0]
        lines.append("| " + " | ".join(header) + " |")

        # Trennzeile
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")

        # Datenzeilen
        for row in normalized_rows[1:]:
            # Sicherstellen dass die Zeilenlänge mit dem Header übereinstimmt
            padded_row = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(padded_row[: len(header)]) + " |")

        markdown_parts.append("\n".join(lines))

    return "\n\n".join(markdown_parts)


def is_scanned_pdf(file_path: Path) -> bool:
    """
    Prüft ob eine PDF-Datei ein eingescanntes Dokument ist.

    Nutzt pdftotext (poppler-utils) um Text zu extrahieren und berechnet den
    Durchschnitt der Zeichen pro Seite. Wenn dieser Durchschnitt unter dem
    konfigurierten Schwellwert (SCAN_THRESHOLD_CHARS) liegt, gilt die PDF
    als Scan.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        True wenn die Datei als Scan erkannt wurde, False sonst.
    """
    try:
        result = subprocess.run(
            ["pdftotext", str(file_path), "-"],
            capture_output=True,
            text=True,
            timeout=PDFTOTEXT_TIMEOUT,
        )
        if result.returncode != 0:
            log.warning(
                "pdftotext_failed",
                file=str(file_path),
                returncode=result.returncode,
                stderr=result.stderr,
            )
            return False

        extracted_text = result.stdout

        # Seitenanzahl ermitteln
        page_count_result = subprocess.run(
            ["pdfinfo", str(file_path)],
            capture_output=True,
            text=True,
            timeout=PDFINFO_TIMEOUT,
        )
        pages = 1
        if page_count_result.returncode == 0:
            for line in page_count_result.stdout.splitlines():
                if line.lower().startswith("pages:"):
                    try:
                        pages = int(line.split(":", 1)[1].strip())
                    except ValueError:
                        pass
                    break

        total_chars = len(extracted_text.strip())
        avg_chars_per_page = total_chars / max(pages, 1)

        is_scan = avg_chars_per_page < SCAN_THRESHOLD_CHARS
        log.info(
            "scan_detection",
            file=str(file_path),
            pages=pages,
            total_chars=total_chars,
            avg_chars_per_page=avg_chars_per_page,
            threshold=SCAN_THRESHOLD_CHARS,
            is_scan=is_scan,
        )
        return is_scan

    except FileNotFoundError:
        log.warning("pdftotext_not_found", file=str(file_path))
        return False
    except subprocess.TimeoutExpired:
        log.warning("pdftotext_timeout", file=str(file_path))
        return False
    except Exception as e:
        log.warning("scan_detection_error", file=str(file_path), error=str(e))
        return False


async def convert_scanned_pdf_ocr3(file_path: Path) -> dict[str, Any]:
    """
    Konvertiert ein gescanntes PDF via Mistral OCR 3 API (/v1/ocr).

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Dict mit folgenden Schlüsseln:
        - success (bool)
        - markdown (str): Zusammengeführter Markdown-Text aller Seiten
        - ocr_model (str): Verwendetes OCR-Modell
        - pages (int): Anzahl verarbeiteter Seiten
        - error_code / error: Nur bei Fehler
    """
    if not MISTRAL_API_KEY:
        return {
            "success": False,
            "error_code": ErrorCode.API_KEY_INVALID,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
        }

    log.info("ocr3_convert_start", file=str(file_path), model=MISTRAL_OCR_MODEL)

    try:
        file_data = file_path.read_bytes()
    except Exception as e:
        log.error("ocr3_read_failed", file=str(file_path), error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"Datei konnte nicht gelesen werden: {str(e)}",
        }

    try:
        result = await call_mistral_ocr_api(file_data, file_path.name)
    except httpx.TimeoutException:
        log.error("ocr3_timeout", timeout=MISTRAL_TIMEOUT)
        return {
            "success": False,
            "error_code": ErrorCode.TIMEOUT,
            "error": f"Mistral OCR API Timeout nach {MISTRAL_TIMEOUT}s",
        }
    except httpx.HTTPStatusError as e:
        error_detail = str(e)
        try:
            error_detail = e.response.json().get("error", {}).get("message", str(e))
        except Exception:
            pass
        log.error("ocr3_api_error", error=error_detail)
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"Mistral OCR API Fehler: {error_detail}",
        }
    except Exception as e:
        log.error("ocr3_exception", error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"OCR API Fehler: {str(e)}",
        }

    try:
        pages = result.get("pages", [])
    except (AttributeError, TypeError) as e:
        log.error("ocr3_invalid_response", error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.API_ERROR,
            "error": f"Ungültige OCR-API-Antwort: {str(e)}",
        }
    if not pages:
        log.warning("ocr3_no_pages", file=str(file_path))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": "OCR API lieferte keine Seiten zurück",
        }

    markdown_parts = []
    for page in pages:
        page_index = page.get("index", 0) + 1
        page_markdown = page.get("markdown", "")
        markdown_parts.append(f"## Seite {page_index}\n\n{page_markdown}")

    combined_markdown = "\n\n".join(markdown_parts)

    log.info("ocr3_convert_complete", file=str(file_path), pages=len(pages), model=MISTRAL_OCR_MODEL)

    return {
        "success": True,
        "markdown": combined_markdown,
        "ocr_model": MISTRAL_OCR_MODEL,
        "pages": len(pages),
    }


async def convert_scanned_pdf(file_path: Path, language: str = "de") -> dict[str, Any]:
    """
    Konvertiert ein eingescanntes PDF zu Markdown.

    Primär: Mistral OCR 3 API (/v1/ocr) — wenn MISTRAL_OCR_ENABLED=true.
    Fallback: Mistral Vision — Seiten werden als Bild gerendert und einzeln analysiert.

    Args:
        file_path: Pfad zur gescannten PDF-Datei.
        language: Sprache für den Vision-Prompt (Standard: "de", nur für Fallback relevant).

    Returns:
        Dict mit folgenden Schlüsseln:
        - success (bool)
        - markdown (str): Zusammengeführter Markdown-Text aller Seiten
        - scanned (bool): Immer True
        - pages_processed (int): Anzahl erfolgreich verarbeiteter Seiten (Vision-Pfad)
        - pages (int): Anzahl Seiten (OCR3-Pfad)
        - tokens_per_page (list[dict]): Token-Verbrauch pro Seite (Vision-Pfad)
        - tokens_total (int): Gesamter Token-Verbrauch (Vision-Pfad)
        - vision_model (str): Genutztes Vision-Modell (Vision-Pfad)
        - ocr_model (str): Genutztes OCR-Modell (OCR3-Pfad)
        - error_code / error: Nur bei Fehler
    """
    # Primärer Pfad: Mistral OCR 3
    if MISTRAL_OCR_ENABLED:
        log.info("scanned_pdf_using_ocr3", file=str(file_path), model=MISTRAL_OCR_MODEL)
        ocr3_result = await convert_scanned_pdf_ocr3(file_path)
        if ocr3_result.get("success"):
            log.info("scanned_pdf_ocr3_success", file=str(file_path), pages=ocr3_result.get("pages", 0))
            return {
                "success": True,
                "markdown": ocr3_result["markdown"],
                "scanned": True,
                "pages": ocr3_result.get("pages", 0),
                "ocr_model": ocr3_result.get("ocr_model", MISTRAL_OCR_MODEL),
            }
        else:
            log.warning(
                "scanned_pdf_ocr3_failed_fallback_to_vision",
                file=str(file_path),
                error=ocr3_result.get("error", "unknown"),
            )
            # Fallback: Vision-Pfad (weiter unten)
    else:
        log.info("scanned_pdf_ocr3_disabled_using_vision", file=str(file_path))

    # Fallback / direkter Pfad: Mistral Vision (pdf2image)
    if not PDF2IMAGE_AVAILABLE:
        log.error("pdf2image_not_available", file=str(file_path))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": "pdf2image ist nicht installiert (pip install pdf2image)",
        }

    if not MISTRAL_API_KEY:
        return {
            "success": False,
            "error_code": ErrorCode.API_KEY_INVALID,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
        }

    log.info("scanned_pdf_convert_start", file=str(file_path))

    try:
        pages = convert_from_path(str(file_path), dpi=PDF_RENDER_DPI)
    except Exception as e:
        log.error("pdf2image_convert_failed", file=str(file_path), error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"PDF-Rendering fehlgeschlagen: {str(e)}",
        }

    vision_prompt = (
        "Extrahiere den gesamten Text aus diesem Scan einer PDF-Seite und gib ihn als Markdown zurück.\n\n"
        "Regeln:\n"
        "- Behalte die Dokumentsprache bei — übersetze NICHT\n"
        "- Überschriften → # ## ### Markdown-Syntax\n"
        "- Tabellen → immer als Markdown-Tabelle mit | Spalte | Spalte | und Trennzeile\n"
        "- Listen → - oder 1. Markdown-Syntax\n"
        "- Fußnoten, Seitenzahlen und Kopfzeilen → kursiv in eckigen Klammern, z.B. *[Seite 3]*\n"
        "- Wenn eine Passage unleserlich ist → schreibe [UNLESERLICH]\n"
        "- Wenn die Seite keine Textinhalte enthält → antworte nur mit: [LEERE SEITE]\n\n"
        "Antworte ausschließlich mit dem Markdown-Text."
    )

    markdown_parts: list[str] = []
    tokens_per_page: list[dict] = []
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    pages_processed = 0
    vision_model_used = MISTRAL_VISION_MODEL

    for page_num, page_image in enumerate(pages, start=1):
        # PIL Image → bytes (PNG)
        img_buffer = io.BytesIO()
        # Resize wenn nötig, um Tokens zu sparen
        if page_image.width > IMAGE_MAX_WIDTH:
            ratio = IMAGE_MAX_WIDTH / page_image.width
            new_height = int(page_image.height * ratio)
            page_image = page_image.resize((IMAGE_MAX_WIDTH, new_height), Image.Resampling.LANCZOS)

        page_image.save(img_buffer, format="PNG")
        image_bytes = img_buffer.getvalue()

        log.info(
            "scanned_pdf_page_start",
            file=str(file_path),
            page=page_num,
            total_pages=len(pages),
            image_size=len(image_bytes),
        )

        vision_result = await analyze_with_mistral_vision(
            image_bytes,
            "image/png",
            vision_prompt,
            language,
        )

        page_token_info = {
            "page": page_num,
            "tokens_prompt": vision_result.get("tokens_prompt", 0),
            "tokens_completion": vision_result.get("tokens_completion", 0),
            "tokens_total": vision_result.get("tokens_total", 0),
            "success": vision_result.get("success", False),
        }
        tokens_per_page.append(page_token_info)

        if vision_result.get("success"):
            pages_processed += 1
            total_prompt_tokens += vision_result.get("tokens_prompt", 0)
            total_completion_tokens += vision_result.get("tokens_completion", 0)
            total_tokens += vision_result.get("tokens_total", 0)
            vision_model_used = vision_result.get("vision_model", MISTRAL_VISION_MODEL)
            markdown_parts.append(f"## Seite {page_num}\n\n{vision_result['markdown']}")
            log.info(
                "scanned_pdf_page_done",
                file=str(file_path),
                page=page_num,
                tokens=vision_result.get("tokens_total", 0),
            )
        else:
            log.warning(
                "scanned_pdf_page_failed",
                file=str(file_path),
                page=page_num,
                error=vision_result.get("error", "unknown"),
            )
            markdown_parts.append(
                f"## Seite {page_num}\n\n*Seite konnte nicht verarbeitet werden: "
                f"{vision_result.get('error', 'Vision-Fehler')}*"
            )

    combined_markdown = "\n\n".join(markdown_parts)

    log.info(
        "scanned_pdf_convert_complete",
        file=str(file_path),
        pages_total=len(pages),
        pages_processed=pages_processed,
        tokens_total=total_tokens,
    )

    return {
        "success": True,
        "markdown": combined_markdown,
        "scanned": True,
        "pages_processed": pages_processed,
        "tokens_per_page": tokens_per_page,
        "tokens_prompt": total_prompt_tokens,
        "tokens_completion": total_completion_tokens,
        "tokens_total": total_tokens,
        "vision_model": vision_model_used,
    }


# =============================================================================
# Code-Block-Erkennung + Sprach-Fences (FR-MKIT-005)
# =============================================================================

# Mapping: Sprachname → Liste von Erkennungs-Patterns (case-sensitive Regex)
_LANGUAGE_PATTERNS: list[tuple[str, list[str]]] = [
    ("python",     [r"\bdef ", r"\bimport ", r"\bclass ", r"if __name__", r"\bprint\(", r"\bself\."]),
    ("javascript", [r"\bfunction ", r"\bconst ", r"\blet ", r"=> ", r"\bconsole\.log"]),
    ("java",       [r"\bpublic class\b", r"\bprivate ", r"\bSystem\.out\b", r"\bvoid "]),
    ("sql",        [r"\bSELECT\b", r"\bFROM\b", r"\bWHERE\b", r"\bINSERT INTO\b"]),
    ("html",       [r"<html", r"<div", r"<body", r"<!DOCTYPE"]),
    ("css",        [r"\{color:", r"\bmargin:", r"\bpadding:", r"\bdisplay:"]),
    ("bash",       [r"#!/bin/", r"\becho ", r"if \[", r"\bfi\b", r"\bdone\b"]),
    ("go",         [r"\bfunc ", r"\bpackage ", r"\bimport \(", r"\bfmt\."]),
    ("rust",       [r"\bfn ", r"\blet mut\b", r"\bimpl ", r"\bpub fn\b"]),
    ("cpp",        [r"#include", r"\bint main\b", r"\bprintf\(", r"\bstd::"]),
]

# Minimum score (number of pattern matches) to identify a language
_MIN_LANG_SCORE = 2


def detect_code_language(text: str) -> str:
    """
    Erkennt die Programmiersprache eines Code-Blocks via Regex-Heuristik.

    Args:
        text: Zu analysierender Text-Block.

    Returns:
        Sprachname in Kleinbuchstaben (z.B. "python", "javascript") oder ""
        wenn die Sprache nicht mit ausreichender Sicherheit erkannt werden kann.
    """
    best_lang = ""
    best_score = 0

    for lang, patterns in _LANGUAGE_PATTERNS:
        score = sum(1 for pat in patterns if re.search(pat, text))
        if score > best_score:
            best_score = score
            best_lang = lang

    return best_lang if best_score >= _MIN_LANG_SCORE else ""


def detect_and_fence_code_blocks(markdown: str) -> str:
    """
    Sucht in Markdown-Text nach nicht-gefenctem Code und wrapp ihn in
    ```language ... ``` Fences.

    Erkennungs-Kriterien:
    1. Indentierte Blöcke (4+ Leerzeichen am Zeilenanfang), mindestens 3 Zeilen.
    2. Blöcke mit mindestens 3 Code-Indikatoren (Klammern, Semikolons, Zuweisungen).

    Bereits vorhandene Fences (``` ... ```) werden nicht erneut gewrappt.

    Args:
        markdown: Markdown-Text nach der Konvertierung.

    Returns:
        Markdown-Text mit gefencten Code-Blöcken.
    """
    if not markdown:
        return markdown

    # Schritt 1: Vorhandene Fences aus dem Text ausblenden, damit wir sie nicht
    # versehentlich als Kandidaten erkennen.
    # Wir ersetzen sie durch Platzhalter und stellen sie am Ende wieder her.
    fenced_blocks: list[str] = []
    fence_pattern = re.compile(r"```[\s\S]*?```", re.MULTILINE)

    def _stash_fence(m: re.Match) -> str:
        idx = len(fenced_blocks)
        fenced_blocks.append(m.group(0))
        return f"\x00FENCE{idx}\x00"

    working = fence_pattern.sub(_stash_fence, markdown)

    # Schritt 2: Kandidaten-Blöcke identifizieren.
    # Ein Block ist eine zusammenhängende Gruppe von Zeilen mit 4+ Spaces Einrückung
    # ODER eine Gruppe von Zeilen mit Code-Indikatoren.
    lines = working.split("\n")
    result_lines: list[str] = []
    i = 0

    while i < len(lines):
        line = lines[i]

        # Sammle zusammenhängende Zeilen mit 4+ Spaces
        if re.match(r"^    ", line) and line.strip():
            block_lines: list[str] = []
            j = i
            while j < len(lines) and (re.match(r"^    ", lines[j]) or lines[j].strip() == ""):
                block_lines.append(lines[j])
                j += 1

            # Trailing leere Zeilen aus dem Block entfernen
            while block_lines and not block_lines[-1].strip():
                block_lines.pop()

            # Mindestens 3 nicht-leere Zeilen
            non_empty = [l for l in block_lines if l.strip()]
            if len(non_empty) >= 3:
                block_text = "\n".join(block_lines)
                lang = detect_code_language(block_text)
                fence_open = f"```{lang}" if lang else "```"
                result_lines.append(fence_open)
                result_lines.extend(block_lines)
                result_lines.append("```")
                log.debug(
                    "code_block_fenced",
                    lines=len(non_empty),
                    language=lang or "unknown",
                )
            else:
                result_lines.extend(block_lines)

            i = j
            continue

        result_lines.append(line)
        i += 1

    working = "\n".join(result_lines)

    # Schritt 3: Platzhalter durch originale Fences ersetzen
    for idx, original in enumerate(fenced_blocks):
        working = working.replace(f"\x00FENCE{idx}\x00", original)

    return working


def convert_excel_enhanced(file_path: Path, show_formulas: bool = False) -> dict[str, Any]:
    """
    Konvertiert eine Excel-Datei (.xlsx/.xls) zu Markdown mit erweiterten Features.

    Features (FR-MKIT-007):
    - AC-007-1: Jedes Worksheet als eigene Sektion (## Sheet: Name)
    - AC-007-2: Charts werden als Datentabellen extrahiert (via openpyxl)
    - AC-007-3: Zellen mit Formeln optional annotiert (z.B. 42 [=SUM(A1:A10)])
    - AC-007-4: Merged Cells werden korrekt aufgelöst

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx oder .xls)
        show_formulas: Wenn True, werden Formel-Annotationen (z.B. [=SUM(A1:A10)]) hinzugefügt.

    Returns:
        Dict mit:
        - success (bool)
        - markdown (str): Konvertierter Markdown-Text
        - sheets_count (int): Anzahl der verarbeiteten Sheets
        - charts_count (int): Anzahl der gefundenen Charts
        - error_code / error: Nur bei Fehler
    """
    if not OPENPYXL_AVAILABLE:
        log.warning("openpyxl_not_available", file=str(file_path))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": "openpyxl ist nicht installiert (pip install openpyxl)",
        }

    log.info("excel_enhanced_convert_start", file=str(file_path), show_formulas=show_formulas)

    try:
        # data_only=False um Formeln zu lesen; ein zweites Mal mit data_only=True für Werte
        wb_formulas = openpyxl.load_workbook(str(file_path), data_only=False)
        wb_values = openpyxl.load_workbook(str(file_path), data_only=True)
    except Exception as exc:
        log.error("excel_open_error", file=str(file_path), error=str(exc))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"Excel-Datei konnte nicht geöffnet werden: {str(exc)}",
        }

    markdown_parts: list[str] = []
    total_charts = 0

    for sheet_name in wb_values.sheetnames:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        sheet_parts: list[str] = [f"## Sheet: {sheet_name}"]

        # AC-007-4: Merged Cells auflösen
        # Baue ein Dict: (row, col) → Wert der Hauptzelle des Merge-Bereichs
        merged_cell_values: dict[tuple[int, int], Any] = {}
        for merge_range in ws_values.merged_cells.ranges:
            # Wert der linken oberen Zelle des Bereichs
            top_left = ws_values.cell(merge_range.min_row, merge_range.min_col)
            top_left_formula = ws_formulas.cell(merge_range.min_row, merge_range.min_col)
            for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
                for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                    merged_cell_values[(row_idx, col_idx)] = (
                        top_left.value,
                        top_left_formula.value,
                    )

        # Zeilen und Spalten ermitteln
        rows = list(ws_values.iter_rows())
        formula_rows = list(ws_formulas.iter_rows())

        # Leeres Sheet graceful behandeln (AC: test_empty_sheet_handled)
        if not rows:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            log.debug("excel_empty_sheet", sheet=sheet_name)
            continue

        # Maximale Spaltenanzahl bestimmen (über alle Zeilen)
        max_cols = max((len(row) for row in rows), default=0)

        if max_cols == 0:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Tabellen-Daten aufbauen
        table_rows: list[list[str]] = []

        for row_idx, (row_cells, formula_cells) in enumerate(
            zip(rows, formula_rows), start=1
        ):
            row_data: list[str] = []
            for col_idx, (cell, formula_cell) in enumerate(
                zip(row_cells, formula_cells), start=1
            ):
                # AC-007-4: Merged Cell Wert holen
                if (row_idx, col_idx) in merged_cell_values:
                    raw_value, raw_formula = merged_cell_values[(row_idx, col_idx)]
                else:
                    raw_value = cell.value
                    raw_formula = formula_cell.value

                # Zellwert als String
                cell_str = "" if raw_value is None else str(raw_value)

                # AC-007-3: Formel-Annotation
                if (
                    show_formulas
                    and raw_formula is not None
                    and isinstance(raw_formula, str)
                    and raw_formula.startswith("=")
                ):
                    cell_str = f"{cell_str} [{raw_formula}]"

                # Pipe-Zeichen im Zellinhalt escapen (würde Tabelle brechen)
                cell_str = cell_str.replace("|", "\\|").replace("\n", " ")
                row_data.append(cell_str)

            # Fehlende Spalten auffüllen
            while len(row_data) < max_cols:
                row_data.append("")

            table_rows.append(row_data)

        # Prüfen ob Sheet komplett leer ist (alle Zellen None)
        all_empty = all(
            cell_str == ""
            for row in table_rows
            for cell_str in row
        )
        if all_empty:
            sheet_parts.append("*Kein Inhalt*")
            markdown_parts.append("\n".join(sheet_parts))
            continue

        # Markdown-Tabelle aufbauen
        if table_rows:
            header = table_rows[0]
            # Header mit generischen Spaltennamen falls leer
            display_header = [
                h if h else get_column_letter(i + 1)
                for i, h in enumerate(header)
            ]
            table_lines: list[str] = []
            table_lines.append("| " + " | ".join(display_header) + " |")
            table_lines.append("| " + " | ".join(["---"] * max_cols) + " |")
            for data_row in table_rows[1:]:
                padded = data_row + [""] * (max_cols - len(data_row))
                table_lines.append("| " + " | ".join(padded[:max_cols]) + " |")
            sheet_parts.append("\n".join(table_lines))

        # AC-007-2: Charts extrahieren
        sheet_charts = getattr(ws_values, "_charts", [])
        if sheet_charts:
            for chart_idx, chart in enumerate(sheet_charts, start=1):
                total_charts += 1
                chart_title = ""
                try:
                    if hasattr(chart, "title") and chart.title is not None:
                        title_obj = chart.title
                        # openpyxl chart.title kann str, Title-Objekt oder None sein
                        if isinstance(title_obj, str):
                            chart_title = title_obj
                        elif hasattr(title_obj, "tx") and title_obj.tx is not None:
                            # Title-Objekt mit tx.rich.p[].r[].t Struktur
                            try:
                                texts = []
                                for para in title_obj.tx.rich.p:
                                    for run in para.r:
                                        if run.t:
                                            texts.append(run.t)
                                chart_title = " ".join(texts)
                            except Exception:
                                chart_title = f"Chart {chart_idx}"
                        else:
                            chart_title = f"Chart {chart_idx}"
                    else:
                        chart_title = f"Chart {chart_idx}"
                except Exception:
                    chart_title = f"Chart {chart_idx}"

                chart_parts: list[str] = [f"### Chart: {chart_title}"]

                # Datenserien extrahieren
                try:
                    series_list = []
                    if hasattr(chart, "series"):
                        for serie in chart.series:
                            serie_title = ""
                            try:
                                if hasattr(serie, "title") and serie.title is not None:
                                    st = serie.title
                                    if hasattr(st, "v") and st.v is not None:
                                        serie_title = str(st.v)
                                    elif hasattr(st, "strRef") and st.strRef is not None:
                                        cache = getattr(st.strRef, "strCache", None)
                                        if cache and hasattr(cache, "pt") and cache.pt:
                                            serie_title = str(cache.pt[0].v)
                            except Exception:
                                pass

                            # Werte aus dem Cache holen
                            values: list[str] = []
                            try:
                                val_ref = None
                                if hasattr(serie, "val"):
                                    val_ref = serie.val
                                elif hasattr(serie, "yVal"):
                                    val_ref = serie.yVal

                                if val_ref is not None:
                                    num_cache = getattr(val_ref, "numRef", None)
                                    if num_cache:
                                        num_data = getattr(num_cache, "numCache", None)
                                        if num_data and hasattr(num_data, "pt"):
                                            values = [str(pt.v) for pt in num_data.pt]
                            except Exception:
                                pass

                            series_list.append({
                                "title": serie_title or f"Serie {len(series_list) + 1}",
                                "values": values,
                            })

                    if series_list:
                        # Tabelle mit Serien als Spalten
                        headers = [s["title"] for s in series_list]
                        chart_parts.append("| " + " | ".join(headers) + " |")
                        chart_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
                        max_vals = max((len(s["values"]) for s in series_list), default=0)
                        for vi in range(max_vals):
                            row_vals = []
                            for s in series_list:
                                row_vals.append(s["values"][vi] if vi < len(s["values"]) else "")
                            chart_parts.append("| " + " | ".join(row_vals) + " |")
                    else:
                        chart_parts.append("*Keine Datenserien gefunden*")

                except Exception as chart_exc:
                    log.warning(
                        "excel_chart_extraction_error",
                        sheet=sheet_name,
                        chart=chart_idx,
                        error=str(chart_exc),
                    )
                    chart_parts.append("*Chart-Daten konnten nicht extrahiert werden*")

                sheet_parts.append("\n".join(chart_parts))

        markdown_parts.append("\n\n".join(sheet_parts))

    wb_values.close()
    wb_formulas.close()

    combined_markdown = "\n\n".join(markdown_parts)
    sheets_count = len(wb_values.sheetnames)

    log.info(
        "excel_enhanced_convert_done",
        file=str(file_path),
        sheets=sheets_count,
        charts=total_charts,
        chars=len(combined_markdown),
    )

    return {
        "success": True,
        "markdown": combined_markdown,
        "sheets_count": sheets_count,
        "charts_count": total_charts,
    }


# =============================================================================
# DOCX Extras: Kommentare, Header/Footer, Track Changes (FR-MKIT-008)
# =============================================================================

def extract_docx_extras(file_path: Path) -> dict:
    """
    Extrahiert erweiterte Metadaten aus einer DOCX-Datei (FR-MKIT-008).

    Verarbeitet:
    - Kommentare aus word/comments.xml (Author, Date, Text)
    - Header und Footer aus allen Dokumentsektionen via python-docx
    - Track Changes (Einfügungen/Löschungen) aus dem DOCX-XML (w:ins, w:del)

    Args:
        file_path: Pfad zur DOCX-Datei.

    Returns:
        Dict mit:
        - comments: Liste von Dicts mit 'author', 'date', 'text'
        - headers: Liste von Header-Texten (nicht leer)
        - footers: Liste von Footer-Texten (nicht leer)
        - track_changes: Liste von Dicts mit 'type' ('insertion'/'deletion'), 'author', 'date', 'text'
    """
    import xml.etree.ElementTree as ET

    result: dict = {
        "comments": [],
        "headers": [],
        "footers": [],
        "track_changes": [],
    }

    W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

    # --- Kommentare aus word/comments.xml ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "word/comments.xml" in zf.namelist():
                xml_data = zf.read("word/comments.xml")
                root = ET.fromstring(xml_data)
                for comment in root.findall(f"{{{W_NS}}}comment"):
                    author = comment.get(f"{{{W_NS}}}author", "")
                    date = comment.get(f"{{{W_NS}}}date", "")
                    # Text aus allen w:t Elementen zusammenführen
                    text_parts = [t.text or "" for t in comment.findall(f".//{{{W_NS}}}t")]
                    text = " ".join(text_parts).strip()
                    if text:
                        result["comments"].append({
                            "author": author,
                            "date": date,
                            "text": text,
                        })
                log.info(
                    "docx_comments_extracted",
                    file=str(file_path),
                    count=len(result["comments"]),
                )
    except Exception as e:
        log.warning("docx_comments_error", file=str(file_path), error=str(e))

    # --- Header/Footer via python-docx ---
    try:
        import docx as _docx  # python-docx
        doc = _docx.Document(str(file_path))
        for section in doc.sections:
            for hf_obj, target_list in [
                (section.header, result["headers"]),
                (section.footer, result["footers"]),
            ]:
                try:
                    text = hf_obj.text.strip() if hf_obj and hasattr(hf_obj, "text") else ""
                    if not text:
                        # Manuell aus Paragraphen extrahieren
                        if hf_obj and hasattr(hf_obj, "paragraphs"):
                            parts = [p.text.strip() for p in hf_obj.paragraphs if p.text.strip()]
                            text = " | ".join(parts)
                    if text and text not in target_list:
                        target_list.append(text)
                except Exception as inner_e:
                    log.debug("docx_hf_paragraph_error", error=str(inner_e))
        log.info(
            "docx_headers_footers_extracted",
            file=str(file_path),
            headers=len(result["headers"]),
            footers=len(result["footers"]),
        )
    except Exception as e:
        log.warning("docx_headers_footers_error", file=str(file_path), error=str(e))

    # --- Track Changes aus word/document.xml (w:ins, w:del) ---
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "word/document.xml" in zf.namelist():
                xml_data = zf.read("word/document.xml")
                root = ET.fromstring(xml_data)

                for tag, change_type in [("ins", "insertion"), ("del", "deletion")]:
                    for elem in root.findall(f".//{{{W_NS}}}{tag}"):
                        author = elem.get(f"{{{W_NS}}}author", "")
                        date = elem.get(f"{{{W_NS}}}date", "")
                        # w:delText für Löschungen, w:t für Einfügungen
                        if change_type == "deletion":
                            text_parts = [
                                t.text or ""
                                for t in elem.findall(f".//{{{W_NS}}}delText")
                            ]
                        else:
                            text_parts = [
                                t.text or ""
                                for t in elem.findall(f".//{{{W_NS}}}t")
                            ]
                        text = "".join(text_parts)
                        if text:
                            result["track_changes"].append({
                                "type": change_type,
                                "author": author,
                                "date": date,
                                "text": text,
                            })

                log.info(
                    "docx_track_changes_extracted",
                    file=str(file_path),
                    count=len(result["track_changes"]),
                )
    except Exception as e:
        log.warning("docx_track_changes_error", file=str(file_path), error=str(e))

    return result


def append_docx_extras_to_markdown(markdown: str, extras: dict) -> str:
    """
    Fügt DOCX-Extras (Kommentare, Header/Footer, Track Changes) als Markdown-Sektionen an.

    Sektionen werden nur angefügt, wenn die jeweiligen Extras nicht leer sind:
    - ## Kommentare → Blockquotes mit Author und Date
    - ## Header und Footer → Inhalt als Liste
    - ## Änderungsverfolgung → Einfügungen und Löschungen als Diff-Notation

    Args:
        markdown: Bereits konvertierter Markdown-Text.
        extras: Rückgabe-Dict von extract_docx_extras().

    Returns:
        Erweiterter Markdown-String.
    """
    sections: list[str] = []

    # --- Kommentare als Blockquotes ---
    comments = extras.get("comments", [])
    if comments:
        lines = ["## Kommentare", ""]
        for c in comments:
            author = c.get("author", "Unbekannt")
            date = c.get("date", "")
            text = c.get("text", "")
            date_str = f" ({date})" if date else ""
            lines.append(f"> **{author}**{date_str}: {text}")
            lines.append("")
        sections.append("\n".join(lines).rstrip())

    # --- Header und Footer ---
    headers = extras.get("headers", [])
    footers = extras.get("footers", [])
    if headers or footers:
        lines = ["## Header und Footer", ""]
        if headers:
            lines.append("**Header:**")
            for h in headers:
                lines.append(f"- {h}")
            lines.append("")
        if footers:
            lines.append("**Footer:**")
            for f in footers:
                lines.append(f"- {f}")
            lines.append("")
        sections.append("\n".join(lines).rstrip())

    # --- Track Changes als Diff-Notation ---
    track_changes = extras.get("track_changes", [])
    if track_changes:
        lines = ["## Änderungsverfolgung", ""]
        lines.append("```diff")
        for tc in track_changes:
            change_type = tc.get("type", "")
            author = tc.get("author", "")
            date = tc.get("date", "")
            text = tc.get("text", "")
            meta = f"  # {author}" if author else ""
            if date:
                meta += f" ({date})" if author else f"  # ({date})"
            if change_type == "insertion":
                lines.append(f"+ {text}{meta}")
            elif change_type == "deletion":
                lines.append(f"- {text}{meta}")
        lines.append("```")
        sections.append("\n".join(lines))

    if not sections:
        return markdown

    return markdown.rstrip() + "\n\n" + "\n\n".join(sections)


# =============================================================================
# FR-MKIT-009: PDF-Metadaten (Bookmarks, Annotationen, Formularfelder)
# =============================================================================

def extract_pdf_metadata(file_path: Path) -> dict[str, Any]:
    """
    Extrahiert PDF-Metadaten mit PyMuPDF (fitz).

    Liefert:
    - toc: Inhaltsverzeichnis/Bookmarks als Liste von [level, title, page].
    - annotations: Annotationen/Kommentare mit type, content, author, page.
    - form_fields: Formularfelder mit field_name, field_value, field_type, page.

    Wenn PyMuPDF nicht installiert ist, werden leere Listen zurückgegeben.

    Args:
        file_path: Pfad zur PDF-Datei.

    Returns:
        Dict mit Schlüsseln "toc", "annotations", "form_fields".
    """
    if not PYMUPDF_AVAILABLE:
        log.debug("pymupdf_not_available", file=str(file_path))
        return {"toc": [], "annotations": [], "form_fields": []}

    toc: list[list[Any]] = []
    annotations: list[dict[str, Any]] = []
    form_fields: list[dict[str, Any]] = []

    try:
        doc = fitz.open(str(file_path))

        # Bookmarks / Table of Contents
        toc = doc.get_toc()  # Returns list of [level, title, page]

        # Annotationen und Formularfelder seitenweise
        for page_num, page in enumerate(doc, start=1):
            # Annotationen
            for annot in page.annots():
                info = annot.info
                annot_entry: dict[str, Any] = {
                    "page": page_num,
                    "type": annot.type[1] if annot.type else "Unknown",
                    "content": info.get("content", ""),
                    "author": info.get("title", ""),
                }
                annotations.append(annot_entry)

            # Formularfelder (Widgets)
            for widget in page.widgets():
                field_entry: dict[str, Any] = {
                    "page": page_num,
                    "field_name": widget.field_name or "",
                    "field_value": str(widget.field_value) if widget.field_value is not None else "",
                    "field_type": widget.field_type_string or str(widget.field_type),
                }
                form_fields.append(field_entry)

        doc.close()
        log.debug(
            "pdf_metadata_extracted",
            file=str(file_path),
            toc_entries=len(toc),
            annotations=len(annotations),
            form_fields=len(form_fields),
        )

    except Exception as exc:
        log.warning("pdf_metadata_error", file=str(file_path), error=str(exc))

    return {"toc": toc, "annotations": annotations, "form_fields": form_fields}


def prepend_pdf_toc(markdown: str, toc: list[list[Any]]) -> str:
    """
    Fügt ein Markdown-Inhaltsverzeichnis aus PDF-Bookmarks VOR dem Inhalt ein.

    Level-Mapping: Level 1 → ##, Level 2 → ###, Level 3 → ####, etc.

    Args:
        markdown: Bestehender Markdown-Text.
        toc: Liste von [level, title, page] aus fitz.Document.get_toc().

    Returns:
        Markdown mit vorangestelltem Inhaltsverzeichnis.
    """
    if not toc:
        return markdown

    lines = ["## Inhaltsverzeichnis", ""]
    for entry in toc:
        if len(entry) >= 3:
            level, title, page = entry[0], entry[1], entry[2]
        elif len(entry) == 2:
            level, title, page = entry[0], entry[1], None
        else:
            continue
        # Level 1 → ##, Level 2 → ###, ...
        prefix = "#" * (level + 1)
        page_str = f" *(Seite {page})*" if page and page > 0 else ""
        lines.append(f"{prefix} {title}{page_str}")

    toc_block = "\n".join(lines)
    return toc_block + "\n\n" + markdown


def append_pdf_annotations(markdown: str, annotations: list[dict[str, Any]]) -> str:
    """
    Hängt PDF-Annotationen als Blockquotes an den Markdown-Text an.

    Format:
    ## Annotationen
    > **Author** (Seite N, Typ): Content

    Args:
        markdown: Bestehender Markdown-Text.
        annotations: Liste von Dicts mit page, type, content, author.

    Returns:
        Markdown mit angehängter Annotationen-Sektion.
    """
    if not annotations:
        return markdown

    lines = ["## Annotationen", ""]
    for ann in annotations:
        author = ann.get("author", "")
        page = ann.get("page", "")
        ann_type = ann.get("type", "")
        content = ann.get("content", "")

        author_str = f"**{author}**" if author else "**Unbekannt**"
        meta_parts = []
        if page:
            meta_parts.append(f"Seite {page}")
        if ann_type:
            meta_parts.append(ann_type)
        meta_str = f" ({', '.join(meta_parts)})" if meta_parts else ""

        lines.append(f"> {author_str}{meta_str}: {content}")
        lines.append("")

    section = "\n".join(lines).rstrip()
    return markdown.rstrip() + "\n\n" + section


def append_pdf_form_fields(markdown: str, form_fields: list[dict[str, Any]]) -> str:
    """
    Hängt PDF-Formularfelder als Key-Value-Tabelle an den Markdown-Text an.

    Format:
    ## Formularfelder
    | Feld | Wert | Typ | Seite |
    |------|------|-----|-------|
    | name | wert | typ | 1     |

    Args:
        markdown: Bestehender Markdown-Text.
        form_fields: Liste von Dicts mit field_name, field_value, field_type, page.

    Returns:
        Markdown mit angehängter Formularfelder-Sektion.
    """
    if not form_fields:
        return markdown

    lines = [
        "## Formularfelder",
        "",
        "| Feld | Wert | Typ | Seite |",
        "|------|------|-----|-------|",
    ]
    for field in form_fields:
        name = field.get("field_name", "")
        value = field.get("field_value", "")
        ftype = field.get("field_type", "")
        page = field.get("page", "")
        lines.append(f"| {name} | {value} | {ftype} | {page} |")

    section = "\n".join(lines)
    return markdown.rstrip() + "\n\n" + section


def convert_with_markitdown(file_path: Path, show_formulas: bool = False) -> dict[str, Any]:
    """
    Konvertiert eine Datei mit MarkItDown.

    Für Excel-Dateien (.xlsx/.xls) wird convert_excel_enhanced() verwendet (FR-MKIT-007):
    - Multi-Sheet Ausgabe, Merged Cells, optionale Formel-Annotationen, Chart-Extraktion.

    Für PDF-Dateien wird zusätzlich pdfplumber für Tabellen-Extraktion genutzt:
    - Wenn pdfplumber Tabellen findet, werden diese als Markdown-Tabellen in den
      MarkItDown-Text integriert.
    - Wenn pdfplumber keine Tabellen findet, wird reiner MarkItDown-Output genutzt.

    Für PDF und DOCX wird nach der Konvertierung Code-Block-Erkennung angewendet
    (FR-MKIT-005): Indentierte Blöcke werden erkannt und in Fences gewrappt.

    Für PDF werden zusätzlich PyMuPDF-Metadaten extrahiert (FR-MKIT-009):
    - Bookmarks/TOC werden als Inhaltsverzeichnis vorangestellt.
    - Annotationen werden als Blockquotes angehängt.
    - Formularfelder werden als Key-Value-Tabelle angehängt.

    Args:
        file_path: Pfad zur Datei.
        show_formulas: Excel-Formeln annotieren (nur für .xlsx/.xls, FR-MKIT-007).
    """
    suffix = file_path.suffix.lower()

    # FR-MKIT-007: Excel → enhanced converter
    if suffix in {".xlsx", ".xls"}:
        return convert_excel_enhanced(file_path, show_formulas=show_formulas)

    try:
        log.info("markitdown_convert", file=str(file_path))
        result = md.convert(str(file_path))
        markdown_text = result.text_content

        # PDF-spezifisch: pdfplumber für Tabellen-Extraktion
        if file_path.suffix.lower() == ".pdf" and PDFPLUMBER_AVAILABLE:
            page_tables = extract_tables_with_pdfplumber(file_path)
            if page_tables:
                merged_tables = merge_cross_page_tables(page_tables)
                table_markdown = tables_to_markdown(merged_tables)
                if table_markdown:
                    log.info(
                        "pdfplumber_tables_integrated",
                        file=str(file_path),
                        table_count=len(merged_tables),
                    )
                    markdown_text = markdown_text + "\n\n## Tabellen\n\n" + table_markdown
            elif IMG2TABLE_AVAILABLE:
                # AC-012-1: img2table als Fallback wenn pdfplumber keine Tabellen findet
                img2table_tables = extract_tables_with_img2table(file_path)
                if img2table_tables:
                    table_markdown = tables_to_markdown(img2table_tables)
                    if table_markdown:
                        log.info(
                            "img2table_fallback_integrated",
                            file=str(file_path),
                            table_count=len(img2table_tables),
                            tables_source="img2table",
                        )
                        markdown_text = markdown_text + "\n\n## Tabellen\n\n" + table_markdown

        # FR-MKIT-005: Code-Block-Erkennung für PDF und DOCX
        suffix = file_path.suffix.lower()
        if suffix in {".pdf", ".docx", ".doc"}:
            markdown_text = detect_and_fence_code_blocks(markdown_text)
            log.debug("code_fencing_applied", file=str(file_path))

        # FR-MKIT-008: DOCX Extras (Kommentare, Header/Footer, Track Changes)
        if suffix == ".docx":
            try:
                extras = extract_docx_extras(file_path)
                markdown_text = append_docx_extras_to_markdown(markdown_text, extras)
                log.debug("docx_extras_appended", file=str(file_path))
            except Exception as extras_err:
                log.warning(
                    "docx_extras_failed",
                    file=str(file_path),
                    error=str(extras_err),
                )

        # FR-MKIT-009: PDF-Metadaten (Bookmarks, Annotationen, Formularfelder)
        if suffix == ".pdf":
            try:
                pdf_meta = extract_pdf_metadata(file_path)
                markdown_text = prepend_pdf_toc(markdown_text, pdf_meta.get("toc", []))
                markdown_text = append_pdf_annotations(markdown_text, pdf_meta.get("annotations", []))
                markdown_text = append_pdf_form_fields(markdown_text, pdf_meta.get("form_fields", []))
                log.debug("pdf_metadata_appended", file=str(file_path))
            except Exception as pdf_meta_err:
                log.warning(
                    "pdf_metadata_failed",
                    file=str(file_path),
                    error=str(pdf_meta_err),
                )

        return {
            "success": True,
            "markdown": markdown_text,
            "title": getattr(result, "title", None),
        }
    except Exception as e:
        log.error("markitdown_error", file=str(file_path), error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"MarkItDown Fehler: {str(e)}"
        }


async def convert_url(url: str) -> dict[str, Any]:
    """Konvertiert eine URL zu Markdown."""
    try:
        log.info("url_convert", url=url)
        result = md.convert_url(url)
        return {
            "success": True,
            "markdown": result.text_content,
            "title": getattr(result, "title", None),
        }
    except Exception as e:
        log.error("url_convert_error", url=url, error=str(e))
        return {
            "success": False,
            "error_code": ErrorCode.CONVERSION_FAILED,
            "error": f"URL-Konvertierung fehlgeschlagen: {str(e)}"
        }


# =============================================================================
# Embedded Image Extraction + Description
# =============================================================================

def extract_images_from_docx(file_path: Path) -> list[dict]:
    """
    Extrahiert eingebettete Bilder aus einer DOCX-Datei.

    DOCX ist intern ein ZIP-Archiv. Bilder liegen in word/media/*.

    Args:
        file_path: Pfad zur DOCX-Datei

    Returns:
        Liste von Dicts mit 'name', 'data' (bytes), 'position_hint' (Index)
    """
    images: list[dict] = []
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            media_files = [
                name for name in zf.namelist()
                if name.startswith("word/media/") and not name.endswith("/")
            ]
            for idx, media_name in enumerate(sorted(media_files)):
                try:
                    data = zf.read(media_name)
                    img = Image.open(io.BytesIO(data))
                    width, height = img.size
                    if width < MIN_IMAGE_SIZE_PX or height < MIN_IMAGE_SIZE_PX:
                        log.debug(
                            "skip_small_image_docx",
                            name=media_name,
                            width=width,
                            height=height,
                        )
                        continue
                    images.append({
                        "name": Path(media_name).name,
                        "data": data,
                        "position_hint": idx,
                    })
                except Exception as e:
                    log.warning("docx_image_read_error", name=media_name, error=str(e))
    except Exception as e:
        log.error("docx_open_error", file=str(file_path), error=str(e))
    log.info("docx_images_extracted", file=str(file_path), count=len(images))
    return images


def extract_images_from_pptx(file_path: Path) -> list[dict]:
    """
    Extrahiert eingebettete Bilder aus einer PPTX-Datei.

    PPTX ist intern ein ZIP-Archiv. Bilder liegen in ppt/media/*.

    Args:
        file_path: Pfad zur PPTX-Datei

    Returns:
        Liste von Dicts mit 'name', 'data' (bytes), 'slide_number' (Index)
    """
    images: list[dict] = []
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            media_files = [
                name for name in zf.namelist()
                if name.startswith("ppt/media/") and not name.endswith("/")
            ]
            for idx, media_name in enumerate(sorted(media_files)):
                try:
                    data = zf.read(media_name)
                    img = Image.open(io.BytesIO(data))
                    width, height = img.size
                    if width < MIN_IMAGE_SIZE_PX or height < MIN_IMAGE_SIZE_PX:
                        log.debug(
                            "skip_small_image_pptx",
                            name=media_name,
                            width=width,
                            height=height,
                        )
                        continue
                    images.append({
                        "name": Path(media_name).name,
                        "data": data,
                        "slide_number": idx + 1,
                    })
                except Exception as e:
                    log.warning("pptx_image_read_error", name=media_name, error=str(e))
    except Exception as e:
        log.error("pptx_open_error", file=str(file_path), error=str(e))
    log.info("pptx_images_extracted", file=str(file_path), count=len(images))
    return images


async def classify_image_type(image_data: bytes, mimetype: str) -> str:
    """
    Klassifiziert ein Bild via Mistral Vision in eine von fünf Kategorien.

    Schickt das Bild an die Vision-API mit einem präzisen Klassifizierungs-Prompt
    und gibt genau eine Kategorie zurück.

    Args:
        image_data: Rohe Bildbytes.
        mimetype: MIME-Typ des Bildes (z.B. 'image/png').

    Returns:
        Einer der Werte: 'photo', 'chart', 'diagram', 'text_scan', 'decorative'.
        Fallback: 'photo' bei Fehler oder unbekannter Antwort.
    """
    prompt = (
        "Classify this image into EXACTLY one category. Reply with ONE word only:\n\n"
        "photo      = photograph of real-world objects, people, places\n"
        "chart      = bar chart, line chart, pie chart, data visualization with axes/values\n"
        "diagram    = flowchart, org chart, mind map, network diagram, UML, architecture diagram\n"
        "text_scan  = image of a document, form, invoice, letter, or any image where text is the primary content\n"
        "decorative = logo, icon, background image, decorative graphic with no information value\n\n"
        "Reply with exactly one of: photo, chart, diagram, text_scan, decorative"
    )
    valid_types = {"photo", "chart", "diagram", "text_scan", "decorative"}

    log.info("classify_image_type_start", size=len(image_data), mimetype=mimetype)

    result = await analyze_with_mistral_vision(image_data, mimetype, prompt, language="en")
    if not result.get("success"):
        log.warning(
            "classify_image_type_failed",
            error=result.get("error"),
            fallback="photo",
        )
        return "photo"

    raw: str = result.get("markdown", "").strip().lower()
    # Extrahiere erstes Wort aus der Antwort (falls Modell mehr zurückgibt)
    first_word = raw.split()[0] if raw.split() else ""
    image_type = first_word if first_word in valid_types else "photo"

    log.info("classify_image_type_result", raw_response=raw, image_type=image_type)
    return image_type


async def convert_diagram_to_mermaid(image_data: bytes, mimetype: str) -> str:
    """
    Konvertiert ein Diagramm/Flowchart-Bild in Mermaid-Syntax via Vision-API.

    Nutzt einen spezialisierten Prompt um Flowcharts und andere Diagramme
    als valide Mermaid-Syntax zu extrahieren.

    Args:
        image_data: Rohe Bildbytes des Diagramms.
        mimetype: MIME-Typ des Bildes (z.B. 'image/png').

    Returns:
        Mermaid-Code-Block (```mermaid ... ```) oder Fallback-Text bei Fehler.
    """
    prompt = (
        "Convert this diagram image into valid Mermaid syntax.\n\n"
        "Choose the appropriate diagram type:\n"
        "- Flowchart/decision tree → graph TD\n"
        "- Sequence diagram → sequenceDiagram\n"
        "- Class diagram → classDiagram\n"
        "- Org chart → graph TD with descriptive node labels\n\n"
        "Rules:\n"
        "- Output ONLY the Mermaid code inside ```mermaid ... ``` fences\n"
        "- Use exact labels from the image — do not invent labels\n"
        "- If the image cannot be represented as Mermaid: output ```mermaid\\ngraph TD\\n    A[Nicht darstellbar]\\n```\n"
        "- No explanations, no text outside the code block"
    )

    log.info("convert_diagram_to_mermaid_start", size=len(image_data), mimetype=mimetype)

    result = await analyze_with_mistral_vision(image_data, mimetype, prompt, language="en")
    if not result.get("success"):
        log.warning(
            "convert_diagram_to_mermaid_failed",
            error=result.get("error"),
        )
        return "[Diagramm-Konvertierung nicht verfügbar]"

    mermaid_output: str = strip_llm_artifacts(result.get("markdown", "").strip())
    log.info("convert_diagram_to_mermaid_success", output_length=len(mermaid_output))
    return mermaid_output


async def extract_chart_data(image_data: bytes, mimetype: str) -> str:
    """
    Extrahiert Daten aus einem Chart/Diagramm-Bild als Markdown-Tabelle.

    Nutzt einen spezialisierten Prompt um Achsenbeschriftungen, Datenpunkte
    und Legenden aus Balken-, Linien- und Kreisdiagrammen zu extrahieren.

    Args:
        image_data: Rohe Bildbytes des Charts.
        mimetype: MIME-Typ des Bildes (z.B. 'image/png').

    Returns:
        Markdown-Tabelle mit den extrahierten Daten oder Fallback-Text bei Fehler.
    """
    prompt = (
        "Extract all data from this chart as a Markdown table.\n\n"
        "Instructions:\n"
        "- For bar/line charts: columns = X-axis label + one column per data series; rows = data points\n"
        "- For pie/donut charts: columns = Category | Value | Percentage\n"
        "- Use exact axis labels and legend entries as column headers\n"
        "- If exact values are not readable, use estimates with ~ prefix (e.g. ~42)\n"
        "- If no chart data found: output only [No chart data found]\n\n"
        "Output ONLY the Markdown table. No explanations, no introductions."
    )

    log.info("extract_chart_data_start", size=len(image_data), mimetype=mimetype)

    result = await analyze_with_mistral_vision(image_data, mimetype, prompt, language="en")
    if not result.get("success"):
        log.warning(
            "extract_chart_data_failed",
            error=result.get("error"),
        )
        return "[Daten-Extraktion nicht verfügbar]"

    table_output: str = strip_llm_artifacts(result.get("markdown", "").strip())
    log.info("extract_chart_data_success", output_length=len(table_output))
    return table_output


async def describe_embedded_images(
    images: list[dict],
    language: str = "de",
) -> list[dict]:
    """
    Beschreibt eine Liste extrahierter Bilder via Mistral Pixtral Vision.

    Bilder werden zuerst klassifiziert (AC-004-1):
    - 'diagram' → Mermaid-Syntax-Konvertierung (AC-004-2)
    - 'chart'   → Datentabellen-Extraktion (AC-004-3)
    - 'photo'   → generische Vision-Beschreibung
    - 'text_scan' → generische Vision-Beschreibung (Text-Extraktion)
    - 'decorative' → wird übersprungen

    Args:
        images: Liste von Dicts mit 'name' und 'data' (bytes) aus extract_images_from_*()
        language: Antwortsprache ('de' oder 'en')

    Returns:
        Liste von Dicts mit 'name', 'description', 'tokens', 'image_type'
    """
    results: list[dict] = []
    for image in images:
        name = image["name"]
        data = image["data"]
        mimetype = detect_mimetype_from_bytes(data) or "image/png"
        log.info("describing_embedded_image", name=name, size=len(data))

        # AC-004-1: Bild klassifizieren
        image_type = await classify_image_type(data, mimetype)
        log.info("image_classified", name=name, image_type=image_type)

        if image_type == "decorative":
            # Dekorative Bilder überspringen
            log.info("skip_decorative_image", name=name)
            continue

        if image_type == "diagram":
            # AC-004-2: Flowcharts/Organigramme → Mermaid
            description = await convert_diagram_to_mermaid(data, mimetype)
            results.append({
                "name": name,
                "description": description,
                "tokens": 0,
                "image_type": image_type,
            })
            continue

        if image_type == "chart":
            # AC-004-3: Balken-/Linien-/Kreisdiagramme → Datentabelle
            description = await extract_chart_data(data, mimetype)
            results.append({
                "name": name,
                "description": description,
                "tokens": 0,
                "image_type": image_type,
            })
            continue

        # 'photo' und 'text_scan': differenzierte Vision-Beschreibung
        if image_type == "text_scan":
            generic_prompt = (
                "Extrahiere den gesamten sichtbaren Text aus diesem Bild. "
                "Gib ihn strukturiert als Markdown wieder. Übersetze nicht."
                if language == "de"
                else "Extract all visible text from this image. "
                     "Return it as structured Markdown. Do not translate."
            )
        else:  # photo oder unbekannt
            generic_prompt = (
                "Beschreibe dieses Bild präzise: was ist zu sehen, relevante Beschriftungen, "
                "erkennbare Objekte und der Gesamtkontext. Format: kurzer Absatz."
                if language == "de"
                else "Describe this image precisely: what is shown, relevant labels, "
                     "recognizable objects and overall context. Format: short paragraph."
            )

        result = await analyze_with_mistral_vision(data, mimetype, generic_prompt, language)
        if result["success"]:
            results.append({
                "name": name,
                "description": result["markdown"],
                "tokens": result.get("tokens_total", 0),
                "image_type": image_type,
            })
        else:
            log.warning(
                "embedded_image_description_failed",
                name=name,
                error=result.get("error"),
            )
            results.append({
                "name": name,
                "description": f"[Bildbeschreibung nicht verfügbar: {result.get('error', 'Unbekannter Fehler')}]",
                "tokens": 0,
                "image_type": image_type,
            })
    return results


def insert_image_descriptions(markdown: str, descriptions: list[dict]) -> str:
    """
    Ersetzt Bild-Platzhalter im Markdown durch Pixtral-Beschreibungen.

    Markitdown erzeugt Platzhalter wie ![image](image1.png) für eingebettete Bilder.
    Diese Funktion ersetzt sie durch einen beschreibenden Blockquote.

    Args:
        markdown: Konvertierter Markdown-Text mit Bild-Platzhaltern
        descriptions: Liste von Dicts mit 'name' und 'description'

    Returns:
        Markdown mit eingefügten Bildbeschreibungen
    """
    # Baue Lookup: Dateiname → Beschreibung
    desc_map: dict[str, str] = {d["name"]: d["description"] for d in descriptions}

    def replace_placeholder(match: re.Match) -> str:
        alt_text = match.group(1)
        img_ref = match.group(2)
        # img_ref kann ein Dateiname oder Pfad sein
        img_name = Path(img_ref).name
        if img_name in desc_map:
            description = desc_map[img_name]
            return f"> **[Bild: {img_name}]** {description}"
        # Kein passender Eintrag → Original behalten
        return match.group(0)

    # Muster: ![alt](ref) — erfasst Bild-Platzhalter
    pattern = r"!\[([^\]]*)\]\(([^)]+)\)"
    return re.sub(pattern, replace_placeholder, markdown)


# =============================================================================
# Dokumenten-Klassifizierung via LLM
# =============================================================================

async def classify_document(
    markdown: str,
    categories: list[str] | None = None,
    language: str = "de",
) -> dict[str, Any]:
    """
    Klassifiziert ein Dokument anhand seines Markdown-Inhalts via Mistral API.

    Args:
        markdown: Konvertierter Markdown-Text des Dokuments.
        categories: Erlaubte Dokumenttypen. Wenn None, werden DEFAULT_CLASSIFY_CATEGORIES
                    verwendet.
        language: Sprache des Prompts ('de' oder 'en').

    Returns:
        Dict mit 'document_type' (str) und 'document_type_confidence' (float 0.0–1.0).
        Bei Fehlern wird {"document_type": "other", "document_type_confidence": 0.0}
        zurückgegeben (graceful degradation).
    """
    if not MISTRAL_API_KEY:
        log.warning("classify_document_no_api_key")
        return {"document_type": "other", "document_type_confidence": 0.0}

    effective_categories = categories if categories is not None else DEFAULT_CLASSIFY_CATEGORIES
    categories_str = ", ".join(effective_categories)

    # Markdown auf maximal CLASSIFY_MAX_CHARS Zeichen kürzen, um Token-Kosten zu begrenzen
    truncated_markdown = markdown[:CLASSIFY_MAX_CHARS] if len(markdown) > CLASSIFY_MAX_CHARS else markdown

    if language == "de":
        system_prompt = "Du bist ein Experte für Dokumentenklassifizierung. Antworte ausschließlich mit validem JSON."
        user_prompt = (
            f"Klassifiziere dieses Dokument. Antworte AUSSCHLIESSLICH mit diesem JSON-Format:\n"
            f'{{\"type\": \"invoice\", \"confidence\": 0.95}}\n\n'
            f"Erlaubte Werte für \"type\": {categories_str}\n"
            f"\"confidence\": Zahl zwischen 0.0 (sehr unsicher) und 1.0 (sehr sicher)\n"
            f"Verwende GENAU einen der erlaubten Typen ohne Abweichung.\n\n"
            f"Dokument:\n{truncated_markdown}"
        )
    else:
        system_prompt = "You are an expert document classifier. Respond exclusively with valid JSON."
        user_prompt = (
            f"Classify this document. Respond EXCLUSIVELY with this JSON format:\n"
            f'{{\"type\": \"invoice\", \"confidence\": 0.95}}\n\n'
            f"Allowed values for \"type\": {categories_str}\n"
            f"\"confidence\": number between 0.0 (very uncertain) and 1.0 (very certain)\n"
            f"Use EXACTLY one of the allowed types without deviation.\n\n"
            f"Document:\n{truncated_markdown}"
        )

    payload = {
        "model": MISTRAL_TEXT_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": CLASSIFY_MAX_TOKENS,
        "temperature": 0.0,
    }

    try:
        log.info("classify_document_start", categories=effective_categories, text_length=len(truncated_markdown))
        result = await call_mistral_vision_api(payload)
        content = result["choices"][0]["message"]["content"].strip()

        # JSON aus der Antwort extrahieren (Modell könnte Markdown-Code-Blöcke liefern)
        json_match = re.search(r"\{[^}]+\}", content, re.DOTALL)
        if not json_match:
            log.warning("classify_document_no_json", raw_content=content)
            return {"document_type": "other", "document_type_confidence": 0.0}

        parsed = json.loads(json_match.group(0))
        doc_type = str(parsed.get("type", "other")).strip()
        confidence = float(parsed.get("confidence", 0.0))

        # Nur erlaubte Typen durchlassen
        if doc_type not in effective_categories:
            log.warning("classify_document_unknown_type", doc_type=doc_type, allowed=effective_categories)
            doc_type = "other"

        # Konfidenz auf gültigen Bereich begrenzen
        confidence = max(0.0, min(1.0, confidence))

        log.info("classify_document_done", document_type=doc_type, confidence=confidence)
        return {"document_type": doc_type, "document_type_confidence": confidence}

    except json.JSONDecodeError as exc:
        log.warning("classify_document_json_error", error=str(exc))
        return {"document_type": "other", "document_type_confidence": 0.0}
    except Exception as exc:
        log.warning("classify_document_api_error", error=str(exc))
        return {"document_type": "other", "document_type_confidence": 0.0}


# =============================================================================
# OCR-Nachkorrektur via LLM
# =============================================================================

async def correct_ocr_text(text: str, language: str = "de") -> dict[str, Any]:
    """
    Korrigiert typische OCR-Fehler in einem Text via Mistral API.

    Sendet den OCR-Text mit einem speziellen Korrektur-Prompt an die Mistral API.
    Typische OCR-Artefakte wie Zeichenverwechslungen (rn→m, 0→O, l→1, fi→fi),
    falsche Worttrennungen und fehlende Leerzeichen werden behoben.
    Inhaltliche Fakten, Zahlen, Namen und Daten werden NICHT verändert.

    Args:
        text: OCR-extrahierter Text (Markdown).
        language: Sprache des Textes ('de' oder 'en').

    Returns:
        Dict mit:
        - corrected_text (str): Korrigierter Text
        - corrections_count (int): Anzahl der Korrekturen
        - tokens (int): Verbrauchte Tokens
        - success (bool): War die Korrektur erfolgreich?
        - error (str, optional): Fehlermeldung bei Misserfolg
    """
    if not MISTRAL_API_KEY:
        log.warning("correct_ocr_text_no_api_key")
        return {
            "success": False,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
            "corrected_text": text,
            "corrections_count": 0,
            "tokens": 0,
        }

    if language == "de":
        system_prompt = (
            "Du bist ein Experte für OCR-Fehlerkorrektur. "
            "Korrigiere ausschließlich offensichtliche OCR-Artefakte, verändere KEINE inhaltlichen Fakten."
        )
        user_prompt = (
            "Korrigiere OCR-Fehler in diesem Markdown-Text.\n\n"
            "Erlaubte Korrekturen (NUR diese):\n"
            "- Zeichen-Verwechslungen: rn→m, 0→O, l→1, fi-Ligaturen, Ü→U etc.\n"
            "- Zusammengeklebte Wörter: 'dasHaus' → 'das Haus'\n"
            "- Falsche Worttrennungen: 'Doku-\\nment' → 'Dokument'\n\n"
            "VERBOTEN:\n"
            "- Inhaltliche Korrekturen (Fakten, Zahlen, Namen)\n"
            "- Änderungen an Markdown-Formatierung (#, *, |, ```)\n"
            "- Umformulierungen\n\n"
            "Antworte mit dem korrigierten Text, dann genau eine abschließende Zeile:\n"
            "<<<CORRECTIONS:N>>>\n"
            "(wobei N die Anzahl der Korrekturen ist)\n\n"
            f"Text:\n{text}"
        )
    else:
        system_prompt = (
            "You are an expert in OCR error correction. "
            "Correct only obvious OCR artifacts, do NOT change any factual content."
        )
        user_prompt = (
            "Correct OCR errors in this Markdown text.\n\n"
            "Allowed corrections (ONLY these):\n"
            "- Character confusions: rn→m, 0→O, l→1, fi-ligatures, etc.\n"
            "- Glued-together words: 'theHouse' → 'the House'\n"
            "- Wrong hyphenation: 'docu-\\nment' → 'document'\n\n"
            "FORBIDDEN:\n"
            "- Content corrections (facts, numbers, names)\n"
            "- Changes to Markdown formatting (#, *, |, ```)\n"
            "- Rephrasing\n\n"
            "Reply with the corrected text, then exactly one closing line:\n"
            "<<<CORRECTIONS:N>>>\n"
            "(where N is the number of corrections)\n\n"
            f"Text:\n{text}"
        )

    payload = {
        "model": MISTRAL_TEXT_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": OCR_CORRECT_MAX_TOKENS,
        "temperature": 0.0,
    }

    try:
        log.info("correct_ocr_text_start", text_length=len(text), language=language)
        result = await call_mistral_vision_api(payload)
        content = result["choices"][0]["message"]["content"]
        usage = result.get("usage", {})
        tokens_total = usage.get("total_tokens", 0)

        # Korrektur-Anzahl aus dem speziellen Marker parsen
        corrections_count = 0
        corrected_text = content

        marker_match = re.search(r"<<<CORRECTIONS:(\d+)>>>\s*$", content.strip(), re.MULTILINE)
        if not marker_match:
            # Fallback: altes Marker-Format unterstützen
            marker_match = re.search(r"---CORRECTIONS:\s*(\d+)\s*$", content.strip(), re.MULTILINE)
        if marker_match:
            corrections_count = int(marker_match.group(1))
            # Marker aus dem Text entfernen
            corrected_text = strip_llm_artifacts(content[: marker_match.start()].rstrip())
        else:
            log.warning("correct_ocr_text_no_corrections_marker", content_tail=content[-100:])
            corrected_text = strip_llm_artifacts(content)

        log.info(
            "correct_ocr_text_done",
            corrections_count=corrections_count,
            tokens=tokens_total,
        )
        return {
            "success": True,
            "corrected_text": corrected_text,
            "corrections_count": corrections_count,
            "tokens": tokens_total,
        }

    except Exception as exc:
        log.warning("correct_ocr_text_api_error", error=str(exc))
        return {
            "success": False,
            "error": str(exc),
            "corrected_text": text,
            "corrections_count": 0,
            "tokens": 0,
        }


# =============================================================================
# Schema-basierte strukturierte Extraktion
# =============================================================================

EXTRACTION_TEMPLATES: dict[str, dict] = {
    "invoice": {
        "type": "object",
        "properties": {
            "invoice_number": {"type": "string"},
            "date": {"type": "string"},
            "vendor": {"type": "string"},
            "total_amount": {"type": "number"},
            "currency": {"type": "string"},
            "line_items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "description": {"type": "string"},
                        "quantity": {"type": "number"},
                        "unit_price": {"type": "number"},
                        "total": {"type": "number"},
                    },
                },
            },
        },
    },
    "cv": {
        "type": "object",
        "properties": {
            "name": {"type": "string"},
            "email": {"type": "string"},
            "phone": {"type": "string"},
            "education": {"type": "array", "items": {"type": "object"}},
            "experience": {"type": "array", "items": {"type": "object"}},
            "skills": {"type": "array", "items": {"type": "string"}},
        },
    },
    "contract": {
        "type": "object",
        "properties": {
            "contract_type": {"type": "string"},
            "parties": {"type": "array", "items": {"type": "string"}},
            "effective_date": {"type": "string"},
            "expiration_date": {"type": "string"},
            "key_terms": {"type": "array", "items": {"type": "string"}},
        },
    },
}


async def extract_structured_data(
    markdown: str,
    schema: dict,
    language: str = "de",
) -> dict:
    """
    Extrahiert strukturierte Daten aus Markdown gemäß einem JSON-Schema via Mistral API.

    Sendet den Markdown-Inhalt zusammen mit dem Schema an die Mistral API.
    Die Antwort wird als JSON geparst und optional gegen das Schema validiert.

    Args:
        markdown: Konvertierter Markdown-Text des Dokuments.
        schema: JSON-Schema für die gewünschten extrahierten Felder.
        language: Sprache des Prompts ('de' oder 'en').

    Returns:
        Dict mit:
        - success (bool)
        - extracted (dict): Extrahierte Daten passend zum Schema
        - tokens (int): Verbrauchte Tokens
        - error (str): Fehlermeldung (nur bei success=False)
    """
    if not MISTRAL_API_KEY:
        log.warning("extract_structured_data_no_api_key")
        return {
            "success": False,
            "error": "MISTRAL_API_KEY nicht konfiguriert",
            "extracted": None,
            "tokens": 0,
        }

    schema_str = json.dumps(schema, ensure_ascii=False)

    if language == "de":
        system_prompt = (
            "Du bist ein Experte für Dokumentenanalyse und Datenextraktion. "
            "Antworte ausschließlich mit validem JSON."
        )
        user_prompt = (
            "Extrahiere strukturierte Daten aus diesem Dokument gemäß dem JSON-Schema.\n\n"
            "Regeln:\n"
            "- Extrahiere NUR Werte die explizit im Dokument stehen — erfinde KEINE Werte\n"
            "- Fehlende Felder: null (niemals raten oder interpolieren)\n"
            "- Arrays: leeres Array [] wenn keine Einträge vorhanden\n"
            "- Zahlen: exakt wie im Dokument (keine Umrechnung, keine Rundung)\n\n"
            "Antworte AUSSCHLIESSLICH mit dem JSON-Objekt. Kein Markdown, keine Erklärungen.\n\n"
            f"Schema:\n{json.dumps(schema, indent=2, ensure_ascii=False)}\n\n"
            f"Dokument:\n{markdown[:EXTRACT_MAX_CHARS]}"
        )
    else:
        system_prompt = (
            "You are an expert in document analysis and data extraction. "
            "Respond exclusively with valid JSON."
        )
        user_prompt = (
            "Extract structured data from this document according to the JSON schema.\n\n"
            "Rules:\n"
            "- Extract ONLY values explicitly stated in the document — do NOT invent values\n"
            "- Missing fields: null (never guess or interpolate)\n"
            "- Arrays: empty array [] if no entries present\n"
            "- Numbers: exactly as in the document (no conversion, no rounding)\n\n"
            "Respond EXCLUSIVELY with the JSON object. No Markdown, no explanations.\n\n"
            f"Schema:\n{json.dumps(schema, indent=2, ensure_ascii=False)}\n\n"
            f"Document:\n{markdown[:EXTRACT_MAX_CHARS]}"
        )

    payload = {
        "model": MISTRAL_TEXT_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": EXTRACT_MAX_TOKENS,
        "temperature": 0.0,
    }

    try:
        log.info(
            "extract_structured_data_start",
            schema_keys=list(schema.get("properties", {}).keys()),
            text_length=len(markdown),
        )
        result = await call_mistral_vision_api(payload)
        content = result["choices"][0]["message"]["content"].strip()
        usage = result.get("usage", {})
        tokens = usage.get("total_tokens", 0)

        # JSON aus der Antwort extrahieren (Modell könnte Markdown-Code-Blöcke liefern)
        json_match = re.search(r"\{[\s\S]*\}", content)
        if not json_match:
            log.warning("extract_structured_data_no_json", raw_content=content[:200])
            return {
                "success": False,
                "error": f"Kein JSON in der API-Antwort gefunden: {content[:100]}",
                "extracted": None,
                "tokens": tokens,
            }

        extracted = json.loads(json_match.group(0))

        # Schema-Validierung (AC-014-7)
        if JSONSCHEMA_AVAILABLE:
            try:
                jsonschema.validate(instance=extracted, schema=schema)
                log.info("extract_structured_data_valid", tokens=tokens)
            except jsonschema.ValidationError as ve:
                log.warning(
                    "extract_structured_data_schema_violation",
                    error=str(ve.message),
                    tokens=tokens,
                )
                # Graceful: trotzdem zurückgeben, aber Warnung im Log
        else:
            log.debug("extract_structured_data_no_jsonschema")

        log.info("extract_structured_data_success", tokens=tokens)
        return {
            "success": True,
            "extracted": extracted,
            "tokens": tokens,
        }

    except json.JSONDecodeError as exc:
        log.warning("extract_structured_data_json_decode_error", error=str(exc))
        return {
            "success": False,
            "error": f"JSON-Parsing fehlgeschlagen: {str(exc)}",
            "extracted": None,
            "tokens": 0,
        }
    except Exception as exc:
        log.error("extract_structured_data_api_error", error=str(exc))
        return {
            "success": False,
            "error": f"API-Fehler bei Extraktion: {str(exc)}",
            "extracted": None,
            "tokens": 0,
        }


# =============================================================================
# Quality Scoring (FR-MKIT-010)
# =============================================================================

def calculate_quality_score(markdown: str, meta: dict) -> dict:
    """
    Berechnet einen Qualitäts-Score für konvertierten Markdown-Text.

    Scoring-Komponenten:
    - Zeichendichte (0-0.3): Verhältnis sinnvoller Zeichen zu Whitespace
    - Wort-Qualität (0-0.3): Anteil erkennbarer Wörter (>= 3 Buchstaben)
    - Struktur-Elemente (0-0.2): Headings, Listen, Tabellen, Code-Blöcke
    - OCR/Vision Confidence (0-0.2): Token-Effizienz als Proxy bei vision_used

    Args:
        markdown: Konvertierter Markdown-Text
        meta: Metadaten-Dictionary (kann vision_used, tokens_* enthalten)

    Returns:
        Dict mit quality_score (0.0-1.0) und quality_grade ('poor'/'fair'/'good'/'excellent')
    """
    if not markdown or not markdown.strip():
        return {"quality_score": 0.0, "quality_grade": "poor"}

    text = markdown.strip()
    total_chars = len(text)

    # --- Komponente 1: Zeichendichte (0-0.3) ---
    # Verhältnis von Nicht-Whitespace zu Gesamt-Zeichen
    non_ws_chars = len(re.sub(r'\s', '', text))
    if total_chars > 0:
        density_ratio = non_ws_chars / total_chars
        # Gute Dichte ist 0.4-0.8; sehr niedrig (<0.2) oder sehr hoch (>0.9) ist suspicious
        if density_ratio < 0.1:
            density_score = 0.0
        elif density_ratio < 0.2:
            density_score = density_ratio * 1.0  # linear bis 0.2
        elif density_ratio <= 0.8:
            density_score = 0.3  # Optimal
        else:
            density_score = max(0.1, 0.3 * (1.0 - density_ratio))
    else:
        density_score = 0.0

    # --- Komponente 2: Wort-Qualität (0-0.3) ---
    # Anteil erkennbarer Wörter (mindestens 3 Buchstaben, keine Gibberish-Sequenzen)
    words = re.findall(r'[a-zA-ZäöüÄÖÜß]{3,}', text)
    total_word_tokens = re.findall(r'\S+', text)

    if total_word_tokens:
        word_ratio = len(words) / len(total_word_tokens)
        word_score = min(0.3, word_ratio * 0.3)
    else:
        word_score = 0.0

    # Bonus: Mindestlänge — sehr kurzer Text bekommt Abzug
    if total_chars < 50:
        word_score *= 0.5

    # --- Komponente 3: Struktur-Elemente (0-0.2) ---
    structure_score = 0.0
    lines = text.split('\n')

    has_headings = any(line.strip().startswith('#') for line in lines)
    has_lists = any(re.match(r'^\s*[-*+]\s', line) or re.match(r'^\s*\d+\.\s', line) for line in lines)
    has_tables = any('|' in line and line.count('|') >= 2 for line in lines)
    has_codeblocks = '```' in text

    structure_elements = sum([has_headings, has_lists, has_tables, has_codeblocks])
    structure_score = min(0.2, structure_elements * 0.05)

    # --- Komponente 4: OCR/Vision Confidence (0-0.2) ---
    vision_score = 0.0
    vision_used = meta.get("vision_used", False)
    scanned = meta.get("scanned", False)

    if vision_used or scanned:
        tokens_prompt = meta.get("tokens_prompt") or 0
        tokens_completion = meta.get("tokens_completion") or 0

        if tokens_prompt > 0 and tokens_completion > 0:
            # Token-Effizienz: mehr Output-Tokens relativ zu Input → guter Inhalt extrahiert
            efficiency = tokens_completion / tokens_prompt
            if efficiency >= 0.5:
                vision_score = 0.2
            elif efficiency >= 0.2:
                vision_score = 0.15
            elif efficiency >= 0.05:
                vision_score = 0.1
            else:
                vision_score = 0.05
        elif tokens_completion > 0:
            # Nur Completion bekannt → Mindest-Score
            vision_score = 0.1
        else:
            # Vision wurde verwendet, aber keine Token-Daten → neutraler Wert
            vision_score = 0.1
    else:
        # Kein Vision → volle 0.2 als Baseline (keine Unsicherheit durch OCR)
        vision_score = 0.2

    # --- Gesamt-Score ---
    raw_score = density_score + word_score + structure_score + vision_score
    quality_score = round(min(1.0, max(0.0, raw_score)), 4)

    # --- Grade Mapping ---
    if quality_score < 0.3:
        quality_grade = "poor"
    elif quality_score < 0.6:
        quality_grade = "fair"
    elif quality_score < 0.8:
        quality_grade = "good"
    else:
        quality_grade = "excellent"

    log.debug(
        "quality_score_calculated",
        score=quality_score,
        grade=quality_grade,
        density=density_score,
        words=word_score,
        structure=structure_score,
        vision=vision_score,
    )

    return {"quality_score": quality_score, "quality_grade": quality_grade}


# =============================================================================
# Smart Chunking (FR-MKIT-011)
# =============================================================================

def chunk_markdown(markdown: str, chunk_size: int = 512, source: str = "") -> list[dict]:
    """
    Splittet Markdown intelligent an Heading-Grenzen für RAG-Anwendungen.

    Algorithmus:
    1. Identifiziert alle Headings (# ## ### etc.) mit ihren Positionen
    2. Schützt "atomare" Blöcke: Tabellen (| ... |) und Code-Blöcke (``` ... ```)
    3. Splittet an Headings; wenn ein Chunk > chunk_size Tokens, werden Absätze
       (doppelte Newlines) als sekundäre Split-Punkte genutzt
    4. Tabellen und Code-Blöcke werden niemals zerstückelt

    Args:
        markdown: Der zu chunkende Markdown-Text
        chunk_size: Maximale Chunk-Größe in Tokens (Heuristik: len(text) / 4)
        source: Quelldatei-Name (wird in Chunk-Metadaten eingebettet)

    Returns:
        Liste von Chunk-Dicts mit: index, heading, source, token_count, text
    """
    if not markdown or not markdown.strip():
        return []

    def _token_count(text: str) -> int:
        """Heuristische Token-Schätzung: len / 4."""
        return max(1, len(text) // 4)

    def _is_atomic_block_start(line: str) -> tuple[bool, str]:
        """Prüft ob eine Zeile den Start eines atomaren Blocks markiert."""
        stripped = line.strip()
        if stripped.startswith("```"):
            return True, "code"
        if stripped.startswith("|") and stripped.endswith("|"):
            return True, "table"
        return False, ""

    def _split_into_sections(text: str) -> list[tuple[str, str]]:
        """
        Teilt Text in (heading, content)-Paare auf.
        Heading ist leer für Inhalt vor dem ersten Heading.
        Atomare Blöcke (Code/Tabellen) werden nicht zerstückelt.
        """
        lines = text.split("\n")
        sections: list[tuple[str, str]] = []
        current_heading = ""
        current_lines: list[str] = []
        in_code_block = False
        in_table = False

        for line in lines:
            stripped = line.strip()

            # Code-Block tracking
            if stripped.startswith("```"):
                if in_code_block:
                    # Ende des Code-Blocks
                    current_lines.append(line)
                    in_code_block = False
                    continue
                else:
                    in_code_block = True
                    current_lines.append(line)
                    continue

            if in_code_block:
                current_lines.append(line)
                continue

            # Tabellen-Tracking (Zeilen die mit | anfangen und enden)
            if stripped.startswith("|") and stripped.endswith("|"):
                in_table = True
                current_lines.append(line)
                continue
            else:
                in_table = False

            # Heading-Erkennung (nur außerhalb atomarer Blöcke)
            heading_match = re.match(r"^(#{1,6})\s+(.+)$", line)
            if heading_match:
                # Speichert bisherigen Abschnitt
                if current_lines or current_heading:
                    sections.append((current_heading, "\n".join(current_lines)))
                current_heading = line.strip()
                current_lines = []
            else:
                current_lines.append(line)

        # Letzten Abschnitt speichern
        if current_lines or current_heading:
            sections.append((current_heading, "\n".join(current_lines)))

        return sections

    def _split_at_paragraphs(heading: str, content: str, chunk_size: int) -> list[tuple[str, str]]:
        """
        Splittet langen Inhalt an Absatz-Grenzen (doppelte Newlines).
        Respektiert Code-Blöcke und Tabellen als atomare Einheiten.
        """
        # Wenn unter Größenlimit: direkt zurückgeben
        full_text = (heading + "\n\n" + content).strip() if heading else content.strip()
        if _token_count(full_text) <= chunk_size:
            return [(heading, content)]

        # Teile an doppelten Newlines auf, aber behalte Code/Tabellen zusammen
        paragraphs: list[str] = []
        current_para: list[str] = []
        in_code = False
        in_table = False

        for line in content.split("\n"):
            stripped = line.strip()

            if stripped.startswith("```"):
                if in_code:
                    current_para.append(line)
                    in_code = False
                else:
                    in_code = True
                    current_para.append(line)
                continue

            if in_code:
                current_para.append(line)
                continue

            if stripped.startswith("|") and stripped.endswith("|"):
                in_table = True
                current_para.append(line)
                continue
            else:
                if in_table and stripped == "":
                    # Ende der Tabelle: speichern als atomaren Absatz
                    paragraphs.append("\n".join(current_para))
                    current_para = []
                    in_table = False
                    continue
                in_table = False

            if stripped == "" and current_para and not in_code and not in_table:
                # Paragraph-Grenze
                paragraphs.append("\n".join(current_para))
                current_para = []
            else:
                current_para.append(line)

        if current_para:
            paragraphs.append("\n".join(current_para))

        # Paragraphen zu Chunks zusammenfassen (so viel wie möglich unter chunk_size)
        result_sections: list[tuple[str, str]] = []
        current_chunk_lines: list[str] = []
        current_tokens = _token_count(heading) if heading else 0

        for para in paragraphs:
            if not para.strip():
                continue
            para_tokens = _token_count(para)

            if current_chunk_lines and (current_tokens + para_tokens > chunk_size):
                # Aktuellen Chunk abschließen
                result_sections.append((heading, "\n\n".join(current_chunk_lines)))
                current_chunk_lines = [para]
                current_tokens = ((_token_count(heading) if heading else 0) + para_tokens)
            else:
                current_chunk_lines.append(para)
                current_tokens += para_tokens

        if current_chunk_lines:
            result_sections.append((heading, "\n\n".join(current_chunk_lines)))

        return result_sections if result_sections else [(heading, content)]

    # Schritt 1: Text in Heading-Sektionen aufteilen
    sections = _split_into_sections(markdown)

    # Schritt 2: Lange Sektionen an Absatz-Grenzen weiter aufteilen
    fine_sections: list[tuple[str, str]] = []
    for heading, content in sections:
        sub_sections = _split_at_paragraphs(heading, content, chunk_size)
        fine_sections.extend(sub_sections)

    # Schritt 3: Chunks mit Metadaten erzeugen
    chunks: list[dict] = []
    for idx, (heading, content) in enumerate(fine_sections):
        # Chunk-Text zusammensetzen
        if heading and content.strip():
            chunk_text = heading + "\n\n" + content.strip()
        elif heading:
            chunk_text = heading
        else:
            chunk_text = content.strip()

        if not chunk_text:
            continue

        chunks.append({
            "index": len(chunks),
            "heading": heading,
            "source": source,
            "token_count": _token_count(chunk_text),
            "text": chunk_text,
        })

    log.debug(
        "chunk_markdown_done",
        source=source,
        chunk_count=len(chunks),
        chunk_size=chunk_size,
        total_tokens=sum(c["token_count"] for c in chunks),
    )

    return chunks


# =============================================================================
# Core Konvertierungs-Logik
# =============================================================================

async def convert_auto(
    file_data: bytes,
    filename: str,
    source: str,
    source_type: str,
    input_meta: dict[str, Any],
    prompt: Optional[str] = None,
    language: str = "de",
    describe_images: bool = False,
    classify: bool = False,
    classify_categories: list[str] | None = None,
    extract_schema: Optional[dict] = None,
    ocr_correct: bool = False,
    show_formulas: bool = False,
    chunk: bool = False,
    chunk_size: int = 512,
    accuracy: str = "standard",
) -> ConvertResponse:
    """
    Intelligente Konvertierung basierend auf Dateityp.

    Args:
        file_data: Rohe Datei-Bytes
        filename: Dateiname (wird für Extension-Erkennung genutzt)
        source: Quell-Pfad oder -Bezeichnung (für Metadaten)
        source_type: 'file', 'base64' oder 'url'
        input_meta: Beliebige Pass-through-Metadaten
        prompt: Optionaler Custom-Prompt für Vision
        language: Antwortsprache ('de' oder 'en')
        describe_images: Eingebettete Bilder in DOCX/PPTX durch Pixtral beschreiben
        classify: Dokumenttyp nach Konvertierung via LLM klassifizieren
        classify_categories: Erlaubte Dokumenttypen (überschreibt DEFAULT_CLASSIFY_CATEGORIES)
        extract_schema: JSON-Schema für strukturierte Daten-Extraktion (AC-014-1)
        ocr_correct: OCR-Nachkorrektur via LLM aktivieren (AC-015-5)
        show_formulas: Excel-Formeln im Output annotieren (FR-MKIT-007)
        chunk: Smart Chunking für RAG aktivieren (FR-MKIT-011)
        chunk_size: Maximale Chunk-Größe in Tokens (Default: 512)
        accuracy: Accuracy-Modus: 'standard' (Default) oder 'high'. High aktiviert
                  automatische OCR-Correction und Dual-Pass Vision-Validierung (T-MKIT-020).
    """
    start_time = time.time()
    ext = get_file_extension(filename)
    mimetype = detect_mimetype_from_bytes(file_data) or get_mimetype(Path(filename))

    meta = {
        **input_meta,
        "source": source,
        "source_type": source_type,
        "format": ext.lstrip("."),
        "size_bytes": len(file_data),
    }

    # Größenprüfung
    if len(file_data) > MAX_FILE_SIZE_BYTES:
        meta["duration_ms"] = int((time.time() - start_time) * 1000)
        return create_error_response(
            ErrorCode.FILE_TOO_LARGE,
            f"Datei zu groß: {len(file_data) / 1024 / 1024:.1f}MB (Max: {MAX_FILE_SIZE_MB}MB)",
            meta=meta
        )

    # T-MKIT-020: Accuracy-Modus immer in Meta dokumentieren
    meta["accuracy_mode"] = accuracy

    # Bild → Vision
    if ext in IMAGE_EXTENSIONS or (mimetype and mimetype.startswith("image/")):
        processed_data, resize_meta = resize_image_if_needed(file_data)
        meta.update(resize_meta)
        meta["vision_used"] = True

        vision_prompt = prompt or (
            "Analysiere dieses Bild und gib den Inhalt als Markdown zurück.\n\n"
            "- Wenn Text sichtbar ist: extrahiere ihn vollständig und strukturiert "
            "(Überschriften, Listen, Tabellen in Markdown-Syntax)\n"
            "- Wenn es ein Diagramm, Chart oder Grafik ist: beschreibe die dargestellten Daten präzise\n"
            "- Wenn es ein Foto ohne Text ist: beschreibe den Bildinhalt in einem kurzen Absatz\n"
            "- Wenn die Bildqualität zu schlecht ist: schreibe [UNLESERLICH]\n\n"
            "Antworte ausschließlich mit dem Markdown-Ergebnis."
        )

        result = await analyze_with_mistral_vision(
            processed_data,
            mimetype or "image/jpeg",
            vision_prompt,
            language
        )

        meta["duration_ms"] = int((time.time() - start_time) * 1000)

        if result["success"]:
            meta["vision_model"] = result.get("vision_model")
            meta["tokens_prompt"] = result.get("tokens_prompt")
            meta["tokens_completion"] = result.get("tokens_completion")
            meta["tokens_total"] = result.get("tokens_total")

            markdown = result["markdown"]
            pipeline_steps: list[str] = ["vision"]

            # AC-015: OCR-Nachkorrektur via LLM (nur wenn aktiviert)
            if ocr_correct:
                log.info("ocr_correct_start", path="vision")
                correction = await correct_ocr_text(markdown, language=language)
                if correction["success"]:
                    markdown = correction["corrected_text"]
                    meta["ocr_corrected"] = True
                    meta["ocr_corrections_count"] = correction["corrections_count"]
                    log.info("ocr_correct_done_vision", corrections=correction["corrections_count"])
                else:
                    log.warning("ocr_correct_failed_vision", error=correction.get("error"))
                    meta["ocr_corrected"] = False

            # T-MKIT-020: High-Accuracy-Pipeline für Bilder
            if accuracy == "high":
                log.info("high_accuracy_image_dual_pass_start", file=filename)
                effective_mimetype = mimetype or "image/jpeg"
                markdown = await dual_pass_validate(
                    markdown=markdown,
                    file_data=processed_data,
                    mimetype=effective_mimetype,
                    language=language,
                )
                pipeline_steps.append("dual_pass_validation")

            meta["pipeline_steps"] = pipeline_steps

            if classify:
                classify_result = await classify_document(markdown, classify_categories, language)
                meta.update(classify_result)
            # AC-010: Quality Scoring
            meta.update(calculate_quality_score(markdown, meta))
            response = create_success_response(markdown, meta=meta)
            # AC-014-2/AC-014-3: Strukturierte Extraktion falls Schema gesetzt
            if extract_schema:
                extraction = await extract_structured_data(markdown, extract_schema, language)
                if extraction["success"]:
                    response.extracted = extraction["extracted"]
                    if accuracy == "high":
                        pipeline_steps_updated = list(meta.get("pipeline_steps", pipeline_steps))
                        if "schema_extraction" not in pipeline_steps_updated:
                            pipeline_steps_updated.append("schema_extraction")
                        meta["pipeline_steps"] = pipeline_steps_updated
                        response.meta = MetaData(**{
                            k: v for k, v in meta.items()
                        })
                else:
                    log.warning("extract_structured_data_failed_vision", error=extraction.get("error"))
            # FR-MKIT-011: Smart Chunking für RAG
            if chunk:
                response.chunks = chunk_markdown(markdown, chunk_size=chunk_size, source=source)
            return response
        else:
            return create_error_response(
                result.get("error_code", ErrorCode.VISION_FAILED),
                result["error"],
                meta=meta
            )

    # Audio/Video → faster-whisper Transkription (FR-MKIT-006)
    elif ext in AUDIO_EXTENSIONS or ext in VIDEO_EXTENSIONS:
        temp_media_path = TEMP_DIR / f"{hashlib.md5(file_data).hexdigest()}_{filename}"
        extracted_wav: Optional[Path] = None
        try:
            temp_media_path.write_bytes(file_data)
            audio_path = temp_media_path

            # Video: erst Audio-Track extrahieren
            if ext in VIDEO_EXTENSIONS:
                log.info("video_audio_extract_start", file=filename)
                try:
                    extracted_wav = extract_audio_from_video(temp_media_path)
                    audio_path = extracted_wav
                except RuntimeError as exc:
                    meta["duration_ms"] = int((time.time() - start_time) * 1000)
                    return create_error_response(
                        ErrorCode.CONVERSION_FAILED,
                        f"Audio-Extraktion fehlgeschlagen: {str(exc)}",
                        meta=meta,
                    )

            # Transkribieren
            transcription = transcribe_audio(audio_path)
            meta["duration_ms"] = int((time.time() - start_time) * 1000)

            if not transcription["success"]:
                return create_error_response(
                    ErrorCode.CONVERSION_FAILED,
                    transcription.get("error", "Transkription fehlgeschlagen"),
                    meta=meta,
                )

            # Meta-Daten setzen (AC-006-6)
            meta["language"] = transcription.get("language", "unknown")
            meta["duration_seconds"] = transcription.get("duration", 0.0)
            meta["whisper_model"] = transcription.get("model_size", WHISPER_MODEL_SIZE)

            transcript_text = transcription.get("text", "")
            markdown = f"# Transkription\n\n{transcript_text}"

            log.info(
                "audio_transcription_done",
                file=filename,
                language=meta["language"],
                duration=meta["duration_seconds"],
                chars=len(transcript_text),
            )

            # AC-010: Quality Scoring
            meta.update(calculate_quality_score(markdown, meta))
            response = create_success_response(markdown, meta=meta)
            # FR-MKIT-011: Smart Chunking für RAG
            if chunk:
                response.chunks = chunk_markdown(markdown, chunk_size=chunk_size, source=source)
            return response

        finally:
            temp_media_path.unlink(missing_ok=True)
            if extracted_wav is not None:
                extracted_wav.unlink(missing_ok=True)

    # Dokument → MarkItDown (mit optionalem Scanned-PDF-Routing)
    elif ext in MARKITDOWN_EXTENSIONS or ext:
        temp_path = TEMP_DIR / f"{hashlib.md5(file_data).hexdigest()}_{filename}"
        try:
            temp_path.write_bytes(file_data)

            # Scanned PDF Detection: VOR dem normalen markitdown-Pfad prüfen
            if ext == ".pdf" and is_scanned_pdf(temp_path):
                log.info("scanned_pdf_detected", file=filename)
                result = await convert_scanned_pdf(temp_path, language=language)
                meta["duration_ms"] = int((time.time() - start_time) * 1000)

                if result["success"]:
                    meta["scanned"] = True
                    meta["vision_model"] = result.get("vision_model")
                    meta["ocr_model"] = result.get("ocr_model")
                    meta["tokens_prompt"] = result.get("tokens_prompt")
                    meta["tokens_completion"] = result.get("tokens_completion")
                    meta["tokens_total"] = result.get("tokens_total")
                    meta["tokens_per_page"] = result.get("tokens_per_page")
                    meta["pages_processed"] = result.get("pages_processed")
                    if result.get("ocr_model"):
                        # OCR3-Pfad: kein Vision
                        meta["vision_used"] = False
                    else:
                        meta["vision_used"] = True

                    scanned_markdown = result["markdown"]
                    scanned_pipeline_steps: list[str] = ["ocr"]

                    # AC-015: OCR-Nachkorrektur via LLM (nur wenn explizit aktiviert ODER accuracy=high)
                    if ocr_correct or accuracy == "high":
                        log.info("ocr_correct_start", path="scanned_pdf", accuracy=accuracy)
                        correction = await correct_ocr_text(scanned_markdown, language=language)
                        if correction["success"]:
                            scanned_markdown = correction["corrected_text"]
                            meta["ocr_corrected"] = True
                            meta["ocr_corrections_count"] = correction["corrections_count"]
                            log.info("ocr_correct_done_scanned_pdf", corrections=correction["corrections_count"])
                            scanned_pipeline_steps.append("ocr_correction")
                        else:
                            log.warning("ocr_correct_failed_scanned_pdf", error=correction.get("error"))
                            meta["ocr_corrected"] = False

                    # T-MKIT-020: High-Accuracy → Dual-Pass Validation für gescannte PDFs
                    if accuracy == "high":
                        log.info("high_accuracy_scanned_pdf_dual_pass_start", file=filename)
                        # Erste Seite als Bild rendern für Dual-Pass Validation
                        if PDF2IMAGE_AVAILABLE:
                            try:
                                pdf_pages = convert_from_path(str(temp_path), dpi=PDF_RENDER_DPI, last_page=1)
                                if pdf_pages:
                                    first_page = pdf_pages[0]
                                    if first_page.width > IMAGE_MAX_WIDTH:
                                        ratio = IMAGE_MAX_WIDTH / first_page.width
                                        new_height = int(first_page.height * ratio)
                                        first_page = first_page.resize(
                                            (IMAGE_MAX_WIDTH, new_height), Image.Resampling.LANCZOS
                                        )
                                    page_buf = io.BytesIO()
                                    first_page.save(page_buf, format="PNG")
                                    page_image_bytes = page_buf.getvalue()
                                    scanned_markdown = await dual_pass_validate(
                                        markdown=scanned_markdown,
                                        file_data=page_image_bytes,
                                        mimetype="image/png",
                                        language=language,
                                    )
                                    scanned_pipeline_steps.append("dual_pass_validation")
                            except Exception as dp_exc:
                                log.warning(
                                    "high_accuracy_scanned_pdf_dual_pass_failed",
                                    file=filename,
                                    error=str(dp_exc),
                                )
                        else:
                            log.warning("high_accuracy_scanned_pdf_dual_pass_skipped_no_pdf2image")

                    meta["pipeline_steps"] = scanned_pipeline_steps

                    if classify:
                        classify_result = await classify_document(scanned_markdown, classify_categories, language)
                        meta.update(classify_result)
                    # AC-010: Quality Scoring
                    meta.update(calculate_quality_score(scanned_markdown, meta))
                    response = create_success_response(scanned_markdown, meta=meta)
                    # AC-014-2/AC-014-3: Strukturierte Extraktion falls Schema gesetzt
                    if extract_schema:
                        extraction = await extract_structured_data(scanned_markdown, extract_schema, language)
                        if extraction["success"]:
                            response.extracted = extraction["extracted"]
                            if accuracy == "high":
                                updated_steps = list(meta.get("pipeline_steps", scanned_pipeline_steps))
                                if "schema_extraction" not in updated_steps:
                                    updated_steps.append("schema_extraction")
                                meta["pipeline_steps"] = updated_steps
                                response.meta = MetaData(**{
                                    k: v for k, v in meta.items()
                                })
                        else:
                            log.warning("extract_structured_data_failed_scanned_pdf", error=extraction.get("error"))
                    # FR-MKIT-011: Smart Chunking für RAG
                    if chunk:
                        response.chunks = chunk_markdown(scanned_markdown, chunk_size=chunk_size, source=source)
                    return response
                else:
                    return create_error_response(
                        result.get("error_code", ErrorCode.CONVERSION_FAILED),
                        result["error"],
                        meta=meta
                    )

            result = convert_with_markitdown(temp_path, show_formulas=show_formulas)
            meta["duration_ms"] = int((time.time() - start_time) * 1000)

            if result["success"]:
                if result.get("title"):
                    meta["title"] = result["title"]

                # FR-MKIT-007: Excel-spezifische Metadaten
                if ext in {".xlsx", ".xls"}:
                    if result.get("sheets_count") is not None:
                        meta["sheets_count"] = result["sheets_count"]
                    if result.get("charts_count") is not None:
                        meta["charts_count"] = result["charts_count"]

                markdown = result["markdown"]
                markitdown_pipeline_steps: list[str] = ["markitdown"]

                # Eingebettete Bilder beschreiben (nur für DOCX/PPTX, nur wenn aktiviert)
                if describe_images and ext in {".docx", ".doc", ".pptx", ".ppt"}:
                    log.info("embedded_images_describe_start", file=filename, ext=ext)
                    if ext in {".docx", ".doc"}:
                        images = extract_images_from_docx(temp_path)
                    else:
                        images = extract_images_from_pptx(temp_path)

                    if images:
                        descriptions = await describe_embedded_images(images, language=language)
                        markdown = insert_image_descriptions(markdown, descriptions)
                        meta["images_described"] = len(descriptions)
                        log.info(
                            "embedded_images_described",
                            file=filename,
                            count=len(descriptions),
                        )

                # T-MKIT-020: High-Accuracy → Dual-Pass Validation für PDFs und Bilder
                if accuracy == "high" and ext == ".pdf":
                    log.info("high_accuracy_pdf_dual_pass_start", file=filename)
                    if PDF2IMAGE_AVAILABLE:
                        try:
                            pdf_pages = convert_from_path(str(temp_path), dpi=PDF_RENDER_DPI, last_page=1)
                            if pdf_pages:
                                first_page = pdf_pages[0]
                                if first_page.width > IMAGE_MAX_WIDTH:
                                    ratio = IMAGE_MAX_WIDTH / first_page.width
                                    new_height = int(first_page.height * ratio)
                                    first_page = first_page.resize(
                                        (IMAGE_MAX_WIDTH, new_height), Image.Resampling.LANCZOS
                                    )
                                page_buf = io.BytesIO()
                                first_page.save(page_buf, format="PNG")
                                page_image_bytes = page_buf.getvalue()
                                markdown = await dual_pass_validate(
                                    markdown=markdown,
                                    file_data=page_image_bytes,
                                    mimetype="image/png",
                                    language=language,
                                )
                                markitdown_pipeline_steps.append("dual_pass_validation")
                        except Exception as dp_exc:
                            log.warning(
                                "high_accuracy_pdf_dual_pass_failed",
                                file=filename,
                                error=str(dp_exc),
                            )
                    else:
                        log.warning("high_accuracy_pdf_dual_pass_skipped_no_pdf2image")

                meta["pipeline_steps"] = markitdown_pipeline_steps

                if classify:
                    classify_result = await classify_document(markdown, classify_categories, language)
                    meta.update(classify_result)

                # AC-010: Quality Scoring
                meta.update(calculate_quality_score(markdown, meta))
                response = create_success_response(markdown, meta=meta)
                # AC-014-2/AC-014-3: Strukturierte Extraktion falls Schema gesetzt
                if extract_schema:
                    extraction = await extract_structured_data(markdown, extract_schema, language)
                    if extraction["success"]:
                        response.extracted = extraction["extracted"]
                        if accuracy == "high":
                            updated_steps = list(meta.get("pipeline_steps", markitdown_pipeline_steps))
                            if "schema_extraction" not in updated_steps:
                                updated_steps.append("schema_extraction")
                            meta["pipeline_steps"] = updated_steps
                            response.meta = MetaData(**{
                                k: v for k, v in meta.items()
                            })
                    else:
                        log.warning("extract_structured_data_failed_markitdown", error=extraction.get("error"))
                # FR-MKIT-011: Smart Chunking für RAG
                if chunk:
                    response.chunks = chunk_markdown(markdown, chunk_size=chunk_size, source=source)
                return response
            else:
                return create_error_response(
                    result.get("error_code", ErrorCode.CONVERSION_FAILED),
                    result["error"],
                    meta=meta
                )
        finally:
            temp_path.unlink(missing_ok=True)

    else:
        meta["duration_ms"] = int((time.time() - start_time) * 1000)
        return create_error_response(
            ErrorCode.UNSUPPORTED_FORMAT,
            f"Nicht unterstütztes Format: {ext}",
            meta=meta
        )


async def convert_folder_contents(
    folder_path: Path,
    input_meta: dict[str, Any],
    language: str = "de",
) -> ConvertResponse:
    """
    Konvertiert alle Dateien in einem Ordner zu einem zusammengeführten Markdown.
    """
    start_time = time.time()
    log.info("folder_convert_start", folder=str(folder_path))

    if not folder_path.exists():
        return create_error_response(
            ErrorCode.FILE_NOT_FOUND,
            f"Ordner nicht gefunden: {folder_path}",
            meta=input_meta
        )

    if not folder_path.is_dir():
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            f"Kein Ordner: {folder_path}",
            meta=input_meta
        )

    # Dateien sammeln
    files = sorted([
        f for f in folder_path.iterdir()
        if f.is_file() and not should_skip_file(f.name)
    ], key=lambda f: f.name.lower())

    if not files:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            f"Keine Dateien im Ordner: {folder_path}",
            meta=input_meta
        )

    # Ergebnisse sammeln
    markdown_parts = []
    file_results = []
    total_tokens = 0
    files_processed = 0
    files_failed = 0

    for file_path in files:
        file_meta = {"filename": file_path.name}

        try:
            file_data = file_path.read_bytes()
            result = await convert_auto(
                file_data=file_data,
                filename=file_path.name,
                source=str(file_path),
                source_type="file",
                input_meta={},
                language=language,
            )

            if result.success:
                files_processed += 1
                markdown_parts.append(f"\n\n## {file_path.name}\n\n{result.markdown}")
                file_meta["success"] = True
                file_meta["size_bytes"] = len(file_data)

                if result.meta:
                    if hasattr(result.meta, 'tokens_total') and result.meta.tokens_total:
                        total_tokens += result.meta.tokens_total
                        file_meta["tokens"] = result.meta.tokens_total
                    if hasattr(result.meta, 'vision_used') and result.meta.vision_used:
                        file_meta["vision_used"] = True
            else:
                files_failed += 1
                file_meta["success"] = False
                file_meta["error"] = result.error.message if result.error else "Unknown error"
                markdown_parts.append(f"\n\n## {file_path.name}\n\n*Konvertierung fehlgeschlagen*")

        except Exception as e:
            files_failed += 1
            file_meta["success"] = False
            file_meta["error"] = str(e)
            log.error("folder_file_error", file=file_path.name, error=str(e))

        file_results.append(file_meta)

    # Zusammenführen
    combined_markdown = f"# {folder_path.name}\n" + "".join(markdown_parts)

    meta = {
        **input_meta,
        "source": str(folder_path),
        "source_type": "folder",
        "files_processed": files_processed,
        "files_failed": files_failed,
        "files_total": len(files),
        "tokens_total": total_tokens,
        "files": file_results,
        "duration_ms": int((time.time() - start_time) * 1000),
    }

    log.info("folder_convert_complete",
             folder=str(folder_path),
             processed=files_processed,
             failed=files_failed,
             duration_ms=meta["duration_ms"])

    return create_success_response(combined_markdown, meta=meta)


# =============================================================================
# REST API Endpoints
# =============================================================================

@app.post("/v1/convert", response_model=ConvertResponse)
async def api_convert(request: ConvertRequest) -> ConvertResponse:
    """
    Konvertiert eine Datei zu Markdown.
    """
    inputs = [request.path, request.base64, request.url]
    if sum(1 for x in inputs if x) != 1:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "Genau einer von 'path', 'base64' oder 'url' muss angegeben werden",
            meta=request.meta
        )

    # Template → Schema Auflösung (AC-014-4)
    effective_schema = request.extract_schema
    if request.template and not effective_schema:
        if request.template not in EXTRACTION_TEMPLATES:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                f"Unbekanntes Template: '{request.template}'. Verfügbar: {list(EXTRACTION_TEMPLATES.keys())}",
                meta=request.meta
            )
        effective_schema = EXTRACTION_TEMPLATES[request.template]

    # Pfad-basiert
    if request.path:
        file_path = resolve_path(request.path)
        if not file_path.exists():
            return create_error_response(
                ErrorCode.FILE_NOT_FOUND,
                f"Datei nicht gefunden: {file_path}",
                meta=request.meta
            )
        file_data = file_path.read_bytes()
        return await convert_auto(
            file_data=file_data,
            filename=file_path.name,
            source=str(file_path),
            source_type="file",
            input_meta=request.meta,
            prompt=request.prompt,
            language=request.language,
            describe_images=request.describe_images,
            classify=request.classify,
            classify_categories=request.classify_categories,
            extract_schema=effective_schema,
            ocr_correct=request.ocr_correct,
            show_formulas=request.show_formulas,
            chunk=request.chunk,
            chunk_size=request.chunk_size,
            accuracy=request.accuracy,
        )

    # Base64
    if request.base64:
        if not request.filename:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                "'filename' ist erforderlich bei Base64-Upload",
                meta=request.meta
            )
        try:
            file_data = base64.b64decode(request.base64)
        except Exception as e:
            return create_error_response(
                ErrorCode.INVALID_BASE64,
                f"Ungültiges Base64: {str(e)}",
                meta=request.meta
            )
        return await convert_auto(
            file_data=file_data,
            filename=request.filename,
            source="base64",
            source_type="base64",
            input_meta=request.meta,
            prompt=request.prompt,
            language=request.language,
            describe_images=request.describe_images,
            classify=request.classify,
            classify_categories=request.classify_categories,
            extract_schema=effective_schema,
            ocr_correct=request.ocr_correct,
            show_formulas=request.show_formulas,
            chunk=request.chunk,
            chunk_size=request.chunk_size,
            accuracy=request.accuracy,
        )

    # URL
    if request.url:
        start_time = time.time()
        result = await convert_url(request.url)
        meta = {
            **request.meta,
            "source": request.url,
            "source_type": "url",
            "duration_ms": int((time.time() - start_time) * 1000),
        }
        if result["success"]:
            if result.get("title"):
                meta["title"] = result["title"]
            if request.classify:
                classify_result = await classify_document(
                    result["markdown"], request.classify_categories, request.language
                )
                meta.update(classify_result)
            response = create_success_response(result["markdown"], meta=meta)
            # AC-014-2/AC-014-3: Strukturierte Extraktion für URL
            if effective_schema:
                extraction = await extract_structured_data(result["markdown"], effective_schema, request.language)
                if extraction["success"]:
                    response.extracted = extraction["extracted"]
                else:
                    log.warning("extract_structured_data_failed_url", error=extraction.get("error"))
            # FR-MKIT-011: Smart Chunking für RAG
            if request.chunk:
                response.chunks = chunk_markdown(result["markdown"], chunk_size=request.chunk_size, source=request.url)
            return response
        else:
            return create_error_response(
                result.get("error_code", ErrorCode.CONVERSION_FAILED),
                result["error"],
                meta=meta
            )

    return create_error_response(
        ErrorCode.INTERNAL_ERROR,
        "Unerwarteter Zustand",
        meta=request.meta
    )


@app.post("/v1/convert/folder", response_model=ConvertResponse)
async def api_convert_folder(request: ConvertFolderRequest) -> ConvertResponse:
    """
    Konvertiert alle Dateien in einem Ordner zu Markdown.
    """
    folder_path = resolve_path(request.path)
    return await convert_folder_contents(
        folder_path=folder_path,
        input_meta=request.meta,
        language="de",
    )


@app.post("/v1/extract", response_model=ConvertResponse)
async def api_extract(request: ExtractRequest) -> ConvertResponse:
    """
    Konvertiert ein Dokument zu Markdown UND extrahiert strukturierte Daten in einem Schritt.

    Kombiniert convert + schema-basierte Extraktion via LLM (AC-014-5).
    """
    inputs = [request.path, request.base64, request.url]
    if sum(1 for x in inputs if x) != 1:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "Genau einer von 'path', 'base64' oder 'url' muss angegeben werden",
            meta=request.meta
        )

    # Template → Schema Auflösung (AC-014-4)
    effective_schema = request.extract_schema
    if request.template and not effective_schema:
        if request.template not in EXTRACTION_TEMPLATES:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                f"Unbekanntes Template: '{request.template}'. Verfügbar: {list(EXTRACTION_TEMPLATES.keys())}",
                meta=request.meta
            )
        effective_schema = EXTRACTION_TEMPLATES[request.template]

    if not effective_schema:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "Entweder 'extract_schema' oder 'template' muss angegeben werden",
            meta=request.meta
        )

    # Konvertierungs-Request aus ExtractRequest aufbauen
    convert_req = ConvertRequest(
        path=request.path,
        base64=request.base64,
        filename=request.filename,
        url=request.url,
        language=request.language,
        meta=request.meta,
        extract_schema=effective_schema,
    )
    return await api_convert(convert_req)


@app.get("/v1/templates", response_model=TemplateResponse)
async def api_templates() -> TemplateResponse:
    """
    Gibt alle vordefinierten Extraktions-Templates zurück (AC-014-4).

    Templates können direkt als 'template' Parameter in /v1/convert oder /v1/extract
    übergeben werden.
    """
    return TemplateResponse(templates=EXTRACTION_TEMPLATES)


@app.post("/v1/analyze", response_model=ConvertResponse)
async def api_analyze(request: AnalyzeRequest) -> ConvertResponse:
    """Analysiert ein Bild explizit mit Mistral Vision."""
    if not request.path and not request.base64:
        return create_error_response(
            ErrorCode.INVALID_INPUT,
            "'path' oder 'base64' muss angegeben werden",
            meta=request.meta
        )

    if request.path:
        file_path = resolve_path(request.path)
        if not file_path.exists():
            return create_error_response(
                ErrorCode.FILE_NOT_FOUND,
                f"Datei nicht gefunden: {file_path}",
                meta=request.meta
            )
        file_data = file_path.read_bytes()
        filename = file_path.name
        source = str(file_path)
        source_type = "file"
    else:
        if not request.filename:
            return create_error_response(
                ErrorCode.INVALID_INPUT,
                "'filename' ist erforderlich bei Base64",
                meta=request.meta
            )
        try:
            file_data = base64.b64decode(request.base64)
        except Exception as e:
            return create_error_response(
                ErrorCode.INVALID_BASE64,
                f"Ungültiges Base64: {str(e)}",
                meta=request.meta
            )
        filename = request.filename
        source = "base64"
        source_type = "base64"

    start_time = time.time()
    processed_data, resize_meta = resize_image_if_needed(file_data)
    mimetype = detect_mimetype_from_bytes(file_data) or get_mimetype(Path(filename))

    result = await analyze_with_mistral_vision(
        processed_data,
        mimetype,
        request.prompt,
        request.language
    )

    meta = {
        **request.meta,
        **resize_meta,
        "source": source,
        "source_type": source_type,
        "format": get_file_extension(filename).lstrip("."),
        "size_bytes": len(file_data),
        "vision_used": True,
        "duration_ms": int((time.time() - start_time) * 1000),
    }

    if result["success"]:
        meta["vision_model"] = result.get("vision_model")
        meta["tokens_prompt"] = result.get("tokens_prompt")
        meta["tokens_completion"] = result.get("tokens_completion")
        meta["tokens_total"] = result.get("tokens_total")
        return create_success_response(result["markdown"], meta=meta)
    else:
        return create_error_response(
            result.get("error_code", ErrorCode.VISION_FAILED),
            result["error"],
            meta=meta
        )


@app.get("/v1/health", response_model=HealthResponse)
async def api_health() -> HealthResponse:
    """Health-Check Endpoint."""
    uptime = int(time.time() - START_TIME)
    mistral_configured = bool(MISTRAL_API_KEY)

    return HealthResponse(
        status="ok",
        version=VERSION,
        meta={
            "mistral_api_configured": mistral_configured,
            "vision_model": MISTRAL_VISION_MODEL,
            "mistral_ocr_configured": mistral_configured and MISTRAL_OCR_ENABLED,
            "ocr_model": MISTRAL_OCR_MODEL if MISTRAL_OCR_ENABLED else None,
            "max_file_size_mb": MAX_FILE_SIZE_MB,
            "image_max_width": IMAGE_MAX_WIDTH,
            "max_retries": MAX_RETRIES,
            "uptime_seconds": uptime,
            "mcp_port": MCP_PORT,
            "rest_port": REST_PORT,
        }
    )


@app.get("/v1/formats")
async def api_formats() -> dict:
    """Listet unterstützte Formate auf."""
    return {
        "markitdown": sorted(MARKITDOWN_EXTENSIONS),
        "vision": sorted(IMAGE_EXTENSIONS),
        "audio": sorted(AUDIO_EXTENSIONS),
        "video": sorted(VIDEO_EXTENSIONS),
        "all": sorted(MARKITDOWN_EXTENSIONS | IMAGE_EXTENSIONS | AUDIO_EXTENSIONS | VIDEO_EXTENSIONS),
    }


# =============================================================================
# MCP Tools
# =============================================================================

@mcp.tool(name="convert")
async def mcp_convert(
    path: Optional[str] = None,
    base64_data: Optional[str] = None,
    filename: Optional[str] = None,
    url: Optional[str] = None,
    meta: Optional[dict] = None,
    accuracy: str = "standard",
    classify: bool = False,
    classify_categories: Optional[list] = None,
    describe_images: bool = False,
    ocr_correct: bool = False,
    show_formulas: bool = False,
    chunk: bool = False,
    chunk_size: int = 512,
    extract_schema: Optional[dict] = None,
    template: Optional[str] = None,
    language: str = "de",
    prompt: Optional[str] = None,
) -> str:
    """
    Konvertiert eine Datei zu Markdown (MCP-Version).

    Args:
        path: Dateipfad im Container.
        base64_data: Base64-kodierte Datei (erfordert filename).
        filename: Dateiname (erforderlich bei base64_data).
        url: URL zu Datei oder Webseite.
        meta: Beliebige Metadaten (werden durchgereicht).
        accuracy: Accuracy-Modus: 'standard' (Default) oder 'high'.
        classify: Dokumenttyp via LLM klassifizieren.
        classify_categories: Erlaubte Dokumenttypen (überschreibt Default).
        describe_images: Eingebettete Bilder in DOCX/PPTX beschreiben.
        ocr_correct: OCR-Nachkorrektur via LLM aktivieren.
        show_formulas: Excel-Formeln im Output annotieren.
        chunk: Smart Chunking für RAG aktivieren.
        chunk_size: Maximale Chunk-Größe in Tokens (Default: 512).
        extract_schema: JSON-Schema für strukturierte Daten-Extraktion.
        template: Vordefinierter Template-Name als Alternative zu extract_schema.
        language: Antwortsprache ('de' oder 'en').
        prompt: Optionaler Custom-Prompt für Vision.
    """
    # Template → Schema Auflösung
    effective_schema = extract_schema
    if template and not effective_schema:
        if template not in EXTRACTION_TEMPLATES:
            return json.dumps({
                "success": False,
                "error": f"Unbekanntes Template: '{template}'. Verfügbar: {list(EXTRACTION_TEMPLATES.keys())}"
            })
        effective_schema = EXTRACTION_TEMPLATES[template]

    if path:
        file_path = resolve_path(path)
        if not file_path.exists():
            return json.dumps({"success": False, "error": f"Datei nicht gefunden: {file_path}"})
        file_data = file_path.read_bytes()
        response = await convert_auto(
            file_data=file_data,
            filename=file_path.name,
            source=str(file_path),
            source_type="file",
            input_meta=meta or {},
            prompt=prompt,
            language=language,
            describe_images=describe_images,
            classify=classify,
            classify_categories=classify_categories,
            extract_schema=effective_schema,
            ocr_correct=ocr_correct,
            show_formulas=show_formulas,
            chunk=chunk,
            chunk_size=chunk_size,
            accuracy=accuracy,
        )
    elif base64_data and filename:
        try:
            file_data = base64.b64decode(base64_data)
        except Exception as e:
            return json.dumps({"success": False, "error": f"Ungültiges Base64: {e}"})
        response = await convert_auto(
            file_data=file_data,
            filename=filename,
            source="base64",
            source_type="base64",
            input_meta=meta or {},
            prompt=prompt,
            language=language,
            describe_images=describe_images,
            classify=classify,
            classify_categories=classify_categories,
            extract_schema=effective_schema,
            ocr_correct=ocr_correct,
            show_formulas=show_formulas,
            chunk=chunk,
            chunk_size=chunk_size,
            accuracy=accuracy,
        )
    elif url:
        result = await convert_url(url)
        if result["success"]:
            url_meta: dict[str, Any] = {"source": url, "source_type": "url"}
            if result.get("title"):
                url_meta["title"] = result["title"]
            if classify:
                classify_result = await classify_document(result["markdown"], classify_categories, language)
                url_meta.update(classify_result)
            response = create_success_response(result["markdown"], meta={**(meta or {}), **url_meta})
            if effective_schema:
                extraction = await extract_structured_data(result["markdown"], effective_schema, language)
                if extraction["success"]:
                    response.extracted = extraction["extracted"]
                else:
                    log.warning("mcp_convert_extract_failed_url", error=extraction.get("error"))
            if chunk:
                response.chunks = chunk_markdown(result["markdown"], chunk_size=chunk_size, source=url)
        else:
            response = create_error_response(result.get("error_code", "ERROR"), result["error"])
    else:
        return json.dumps({"success": False, "error": "path, url oder (base64_data + filename) erforderlich"})

    return response.model_dump_json()


@mcp.tool(name="extract")
async def mcp_extract(
    extract_schema: Optional[dict] = None,
    path: Optional[str] = None,
    base64_data: Optional[str] = None,
    filename: Optional[str] = None,
    url: Optional[str] = None,
    template: Optional[str] = None,
    language: str = "de",
    meta: Optional[dict] = None,
    accuracy: str = "standard",
    ocr_correct: bool = False,
    classify: bool = False,
) -> str:
    """
    Konvertiert eine Datei zu Markdown UND extrahiert strukturierte Daten gemäß Schema (AC-014-6).

    Kombiniert convert + schema-basierte Extraktion via Mistral LLM in einem Schritt.
    Die Response enthält sowohl 'markdown' als auch 'extracted' (JSON passend zum Schema).

    Args:
        extract_schema: JSON-Schema für die gewünschten extrahierten Felder (optional wenn template gesetzt).
        path: Dateipfad im Container (alternativ zu base64_data oder url).
        base64_data: Base64-kodierte Datei (erfordert filename).
        filename: Dateiname (erforderlich bei base64_data).
        url: URL zu Datei oder Webseite (alternativ zu path/base64_data).
        template: Vordefinierter Template-Name ('invoice', 'cv', 'contract') als
                  Alternative zu extract_schema.
        language: Antwortsprache ('de' oder 'en').
        meta: Beliebige Metadaten (werden durchgereicht).
        accuracy: Accuracy-Modus: 'standard' (Default) oder 'high'.
        ocr_correct: OCR-Nachkorrektur via LLM aktivieren.
        classify: Dokumenttyp via LLM klassifizieren.
    """
    # Template → Schema Auflösung
    effective_schema = extract_schema
    if template and not effective_schema:
        if template not in EXTRACTION_TEMPLATES:
            return json.dumps({
                "success": False,
                "error": f"Unbekanntes Template: '{template}'. Verfügbar: {list(EXTRACTION_TEMPLATES.keys())}"
            })
        effective_schema = EXTRACTION_TEMPLATES[template]

    if not effective_schema:
        return json.dumps({
            "success": False,
            "error": "Entweder 'extract_schema' oder 'template' muss angegeben werden"
        })

    if path:
        file_path = resolve_path(path)
        if not file_path.exists():
            return json.dumps({"success": False, "error": f"Datei nicht gefunden: {file_path}"})
        file_data = file_path.read_bytes()
        response = await convert_auto(
            file_data=file_data,
            filename=file_path.name,
            source=str(file_path),
            source_type="file",
            input_meta=meta or {},
            language=language,
            extract_schema=effective_schema,
            accuracy=accuracy,
            ocr_correct=ocr_correct,
            classify=classify,
        )
    elif base64_data and filename:
        try:
            file_data = base64.b64decode(base64_data)
        except Exception as e:
            return json.dumps({"success": False, "error": f"Ungültiges Base64: {e}"})
        response = await convert_auto(
            file_data=file_data,
            filename=filename,
            source="base64",
            source_type="base64",
            input_meta=meta or {},
            language=language,
            extract_schema=effective_schema,
            accuracy=accuracy,
            ocr_correct=ocr_correct,
            classify=classify,
        )
    elif url:
        result = await convert_url(url)
        if result["success"]:
            url_meta: dict[str, Any] = {"source": url, "source_type": "url"}
            if classify:
                classify_result = await classify_document(result["markdown"], None, language)
                url_meta.update(classify_result)
            response = create_success_response(result["markdown"], meta={**(meta or {}), **url_meta})
            extraction = await extract_structured_data(result["markdown"], effective_schema, language)
            if extraction["success"]:
                response.extracted = extraction["extracted"]
        else:
            response = create_error_response(result.get("error_code", "ERROR"), result["error"])
    else:
        return json.dumps({"success": False, "error": "path, url oder (base64_data + filename) erforderlich"})

    return response.model_dump_json()


@mcp.tool(name="convert_folder")
async def mcp_convert_folder(
    path: str,
    meta: Optional[dict] = None,
    language: str = "de",
) -> str:
    """
    Konvertiert alle Dateien in einem Ordner zu Markdown.

    Args:
        path: Ordnerpfad im Container.
        meta: Beliebige Metadaten (werden durchgereicht).
        language: Antwortsprache ('de' oder 'en').
    """
    folder_path = resolve_path(path)
    response = await convert_folder_contents(
        folder_path=folder_path,
        input_meta=meta or {},
        language=language,
    )
    return response.model_dump_json()


@mcp.tool(name="health")
async def mcp_health() -> str:
    """Health-Check (MCP-Version)."""
    response = await api_health()
    return response.model_dump_json()


@mcp.tool(name="list_files")
async def mcp_list_files(subdir: str = "") -> str:
    """Listet Dateien im /data Verzeichnis auf."""
    target_dir = DATA_DIR / subdir if subdir else DATA_DIR
    if not target_dir.exists():
        return json.dumps({"error": f"Verzeichnis nicht gefunden: {target_dir}"})

    files = []
    for item in sorted(target_dir.iterdir()):
        if item.is_file():
            files.append({
                "name": item.name,
                "size": item.stat().st_size,
                "type": item.suffix.lower()
            })
        elif item.is_dir():
            files.append({"name": item.name + "/", "type": "directory"})

    return json.dumps({"path": str(target_dir), "files": files}, ensure_ascii=False)


# =============================================================================
# Server Start
# =============================================================================

def run_rest_server():
    """Startet den REST-Server in einem separaten Thread."""
    uvicorn.run(
        app,
        host=os.getenv("BIND_HOST", "0.0.0.0"),
        port=REST_PORT,
        log_level=LOG_LEVEL.lower(),
    )


if __name__ == "__main__":
    log.info("server_starting",
             version=VERSION,
             data_dir=str(DATA_DIR),
             vision_model=MISTRAL_VISION_MODEL,
             vision_enabled=bool(MISTRAL_API_KEY),
             mcp_port=MCP_PORT,
             rest_port=REST_PORT,
             max_file_size_mb=MAX_FILE_SIZE_MB,
             image_max_width=IMAGE_MAX_WIDTH,
             max_retries=MAX_RETRIES)

    print(f"MarkItDown Server v{VERSION}")
    print(f"Data-Verzeichnis: {DATA_DIR}")
    print(f"Vision-Modell: {MISTRAL_VISION_MODEL}")
    print(f"Vision aktiviert: {'Ja' if MISTRAL_API_KEY else 'Nein'}")
    print(f"MCP-Port: {MCP_PORT}")
    print(f"REST-Port: {REST_PORT}")
    print(f"Max Dateigröße: {MAX_FILE_SIZE_MB}MB")
    print(f"Max Bildbreite: {IMAGE_MAX_WIDTH}px")
    print(f"Max Retries: {MAX_RETRIES}")
    print("-" * 50)

    rest_thread = threading.Thread(target=run_rest_server, daemon=True)
    rest_thread.start()
    print(f"REST-API gestartet auf Port {REST_PORT}")

    transport = os.getenv("MCP_TRANSPORT", "sse")
    if transport == "stdio":
        mcp.run()
    else:
        print(f"MCP-Server startet auf Port {MCP_PORT}")
        mcp.run(transport="sse", host=os.getenv("BIND_HOST", "0.0.0.0"), port=MCP_PORT)
